"""
Excel ファイルの軽量バリデーション（アップロード時チェック用）
openpyxl で必須シート・キーワードの存在を確認する。
"""
import unicodedata
from pathlib import Path

import openpyxl


def validate_excel(file_path: Path) -> tuple[bool, str]:
    """Excel 依頼書の基本的な整合性を検証する。

    Returns
    -------
    (valid, message) : tuple[bool, str]
        valid=True なら message は空文字列。
        valid=False なら message にエラー理由。
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as e:
        return False, f"Excelファイルを開けません: {e}"

    try:
        # シート名の部分一致検索（「注文書作成依頼書」を含むシート）
        target_sheet = None
        for name in wb.sheetnames:
            if "注文書作成依頼書" in name:
                target_sheet = wb[name]
                break

        if target_sheet is None:
            return False, "シート「注文書作成依頼書」が見つかりません。"

        # 必須キーワードの存在確認（A列・B列の先頭50行をスキャン）
        required_keywords = ["【工事名】", "【業者名】"]
        found = set()
        for row in target_sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=2, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                text = unicodedata.normalize("NFKC", str(cell)).replace(" ", "").replace("\u3000", "")
                for kw in required_keywords:
                    normalized_kw = unicodedata.normalize("NFKC", kw).replace(" ", "").replace("\u3000", "")
                    if normalized_kw in text:
                        found.add(kw)

        missing = [kw for kw in required_keywords if kw not in found]
        if missing:
            return False, f"必須キーワードが見つかりません: {', '.join(missing)}"

        return True, ""

    finally:
        wb.close()
