"""
Excel データ抽出モジュール — メインオーケストレータ

依頼書 (.xlsx) のメインシートからキーワード検索方式で各業者のデータを集約し、
契約条件書からの共通項目および排他的シート割り当てを統合した結果を返す。

固定セル番地を使わず A列/B列 のキーワードで行番号を動的に特定するため、
依頼書の行がユーザーにより挿入・削除されても追従できる。

抽出ロジックは責務別に下記モジュールへ分散し、ここでは extract_data から
直接 import して呼び出す:
  - extractor_utils         : _cell_str / _cell_raw / 日付パーサ / _safe_int
  - irai_scan_utils         : 依頼書キーワード走査 + 金額抽出
  - sheet_assignment_utils  : 業者×シートのマッチングと共通データ抽出
"""
from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from . import config
from .extractor_utils import (
    _cell_raw,
    _cell_str,
    _parse_contract_date,
    _parse_kouki,
    _safe_int,
)
from .irai_scan_utils import (
    _detect_vendor_base_cols,
    _extract_kingaku_direct,
    _find_sub_keyword_row,
    _scan_keyword_rows,
)
from .sheet_assignment_utils import (
    _extract_from_first_joken,
    build_sheet_assignment,
)

logger = logging.getLogger(__name__)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  メイン抽出関数
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def extract_data(excel_path: Path) -> list[dict[str, str | None]]:
    """
    注文書作成依頼書 (.xlsx) から業者ごとのデータを抽出する。

    キーワード検索方式により、行がユーザーによって挿入・削除されても
    正しいデータ位置を動的に特定できる。

    Parameters
    ----------
    excel_path : Path
        読み込む Excel ファイルのパス。

    Returns
    -------
    list[dict[str, str | None]]
        業者ごとの辞書リスト。各辞書には共通項目・業者固有項目・
        契約日(令和)・工期・金額・対応シート名が含まれる。
        業者が 0 社の場合は空リスト。

    Raises
    ------
    FileNotFoundError
        ファイルが存在しない場合。
    ValueError
        対象シートが見つからない場合。
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel ファイルが見つかりません: {excel_path}")

    logger.info("Excel 読み込み開始: %s", excel_path.name)

    excel_map = config.EXCEL_MAP
    target_sheet_kw = excel_map["sheet_name"]
    common_keywords = excel_map["common_keywords"]
    vendor_keywords = excel_map["vendor_keywords"]
    fallback_base_cols: list[int] = excel_map["vendor_base_cols"]
    excel_map.get("kingaku_sub_keywords", {})

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    try:
        all_sheet_names: list[str] = wb.sheetnames

        # ── 1) 対象シートを特定（部分一致） ──
        ws: Worksheet | None = None
        for name in all_sheet_names:
            if target_sheet_kw in name:
                ws = wb[name]
                logger.info("対象シート特定: '%s'", name)
                break

        if ws is None:
            raise ValueError(
                f"シート名に '{target_sheet_kw}' を含むシートが見つかりません。"
                f" 存在するシート: {all_sheet_names}"
            )

        # シートの有効行数を取得（スキャン範囲）
        max_row = ws.max_row or 200
        if max_row < 10:
            max_row = 200  # 安全マージン

        # ── 1b) 業者基準列を自動検出 ──
        detected_cols = _detect_vendor_base_cols(ws, max_row)
        if detected_cols:
            vendor_base_cols = detected_cols
        else:
            vendor_base_cols = fallback_base_cols
            logger.warning(
                "業者基準列の自動検出に失敗 → config フォールバック値を使用: %s",
                fallback_base_cols,
            )

        # ── 2) キーワードで行番号を特定 ──
        all_keywords = {}
        all_keywords.update(common_keywords)
        all_keywords.update(vendor_keywords)

        keyword_rows = _scan_keyword_rows(ws, all_keywords, max_row)
        logger.info("キーワード検出結果: %s", {k: v for k, v in keyword_rows.items()})

        # ── 3) 共通データ取得（見つかった行の最初の業者基準列から取得） ──
        common_col = vendor_base_cols[0] if vendor_base_cols else 3
        common_data: dict[str, str | None] = {}
        for field_name in common_keywords:
            row = keyword_rows.get(field_name)
            if row is not None:
                common_data[field_name] = _cell_str(ws, row, common_col)
            else:
                common_data[field_name] = None
                logger.warning("共通項目 '%s' の行が未検出", field_name)

        logger.info("共通データ: %s", common_data)

        # ── 4) 契約日ヘッダの下にある「当初」行・「変更1」行を特定 ──
        contract_date_row: int | None = None
        contract_change_row: int | None = None
        contract_header_row = keyword_rows.get("contract_date_header")
        if contract_header_row:
            # 「当初」を探す（ヘッダの 1〜5 行下）
            tosho_row = _find_sub_keyword_row(ws, contract_header_row + 1, "当初", max_search=5)
            if tosho_row:
                contract_date_row = tosho_row
                logger.info("契約日「当初」行: %d", tosho_row)
                # 「変更」行を「当初」の直下で検出
                change_row = _find_sub_keyword_row(ws, tosho_row + 1, "変更", max_search=3)
                if change_row:
                    contract_change_row = change_row
                    logger.info("契約日「変更1」行: %d", change_row)
            else:
                # 「当初」が無い場合はヘッダの直下を使う
                contract_date_row = contract_header_row + 1
                logger.warning("契約日「当初」未検出 → ヘッダ+1行（%d行目）を使用", contract_date_row)

        # ── 5) 工期ヘッダの下にある「当初」行・「変更1」行を特定 ──
        kouki_row: int | None = None
        kouki_change_row: int | None = None
        kouki_header_row = keyword_rows.get("kouki_header")
        if kouki_header_row:
            tosho_row = _find_sub_keyword_row(ws, kouki_header_row + 1, "当初", max_search=5)
            if tosho_row:
                kouki_row = tosho_row
                logger.info("工期「当初」行: %d", tosho_row)
                # 「変更」行を「当初」の直下で検出
                change_row = _find_sub_keyword_row(ws, tosho_row + 1, "変更", max_search=3)
                if change_row:
                    kouki_change_row = change_row
                    logger.info("工期「変更1」行: %d", change_row)
            else:
                kouki_row = kouki_header_row + 1
                logger.warning("工期「当初」未検出 → ヘッダ+1行（%d行目）を使用", kouki_row)

        # ── 6) 金額ヘッダの下にある「当初」行・「変更1」行を特定 ──
        kingaku_tosho_row: int | None = None
        kingaku_change_row: int | None = None
        kingaku_header_row = keyword_rows.get("kingaku_header")
        if kingaku_header_row:
            tosho_row = _find_sub_keyword_row(ws, kingaku_header_row + 1, "当初", max_search=5)
            if tosho_row:
                kingaku_tosho_row = tosho_row
                logger.info("金額「当初」行: %d", tosho_row)
                change_row = _find_sub_keyword_row(ws, tosho_row + 1, "変更", max_search=5)
                if change_row:
                    kingaku_change_row = change_row
                    logger.info("金額「変更1」行: %d", change_row)
            else:
                kingaku_tosho_row = kingaku_header_row + 1
                logger.warning("金額「当初」未検出 → ヘッダ+1行（%d行目）を使用", kingaku_tosho_row)

        # ── 7) 業者名を先行収集（シート一括割り当て用） ──
        # 空欄列はスキップし、途中に空きがあっても後方の業者を拾う
        company_row = keyword_rows.get("vendor_company")
        vendor_companies: list[str] = []
        vendor_active_cols: list[int] = []  # 実際に業者がいる列番号
        if company_row:
            for base_col in vendor_base_cols:
                company = _cell_str(ws, company_row, base_col)
                if company is None or company.strip() == "":
                    continue  # 空欄列はスキップ
                vendor_companies.append(company)
                vendor_active_cols.append(base_col)

        # ── 8) シート一括排他割り当て ──
        # 全業者分を一度に処理し、1シート=1業者の排他制御を保証
        sheet_assignment = build_sheet_assignment(
            all_sheet_names, vendor_companies, wb=wb,
        )
        logger.info("シート割当結果: %s", {
            vc: {k: v for k, v in sheets.items() if v}
            for vc, sheets in sheet_assignment.items()
        })

        # ── 8b) 契約条件書から共通データを抽出 ──
        # 元請負人名称・代表構成員は全業者共通なので、最初の契約条件書から1回だけ取得
        joken_common = _extract_from_first_joken(wb, sheet_assignment)
        if joken_common.get("motouke_company"):
            logger.info("元請負人名称: '%s'", joken_common["motouke_company"])
        else:
            logger.warning("元請負人名称（商号又は名称）が取得できませんでした")
        if joken_common.get("daikyo_koseiin"):
            logger.info("代表構成員: '%s'", joken_common["daikyo_koseiin"])
        else:
            logger.warning("代表構成員が取得できませんでした")

        # ── 9) 業者ループ（列単位） ──
        results: list[dict[str, str | None]] = []

        if not vendor_active_cols:
            if company_row is None:
                logger.warning("【業者名】キーワード未検出 → 業者走査中断")
            else:
                logger.info("業者名が1社も検出できませんでした")

        for col_idx, base_col in enumerate(vendor_active_cols):
            company = vendor_companies[col_idx]
            logger.info("業者 %d: '%s' (基準列=%d)", col_idx + 1, company, base_col)

            vendor_data: dict[str, str | None] = {}
            vendor_data.update(common_data)  # 共通データをマージ
            vendor_data.update(joken_common)  # 契約条件書の共通データをマージ
            vendor_data["vendor_company"] = company

            # ── 代表者名 ──
            name_row = keyword_rows.get("vendor_name")
            if name_row:
                vendor_data["vendor_name"] = _cell_str(ws, name_row, base_col)

            # ── 住所 ──
            addr_row = keyword_rows.get("vendor_address")
            if addr_row:
                vendor_data["vendor_address"] = _cell_str(ws, addr_row, base_col)

            # ── 変更回数 ──
            henkou_row = keyword_rows.get("henkou_kaisuu")
            kaisuu = 0
            if henkou_row:
                # 大見出し行(40行目など)から下方向にスキャンし、各業者の列にある「変更回数」ラベルを探す
                target_row = henkou_row
                for r in range(henkou_row, henkou_row + 5):
                    val = _cell_str(ws, r, base_col)
                    if val and "変更回数" in val:
                        target_row = r
                        break
                
                # 見つけた行の右隣 (base_col + 1) から値を取得 (例: C列がラベルならD列が値)
                raw_kaisuu = _cell_raw(ws, target_row, base_col + 1)
                
                # もし右隣が空なら、元の列も念のため確認
                if raw_kaisuu in (None, "", 0, "0"):
                    raw_kaisuu = _cell_raw(ws, target_row, base_col)
                
                # 数値の抽出（1, 1.0, "第1回" などの表記揺れにすべて対応）
                raw_str = str(raw_kaisuu).strip()
                if raw_str and raw_str not in ("None", "0"):
                    m = re.search(r'\d+', raw_str)
                    if m:
                        kaisuu = int(m.group())
            
            if kaisuu >= 1:
                vendor_data["henkou_kaisuu"] = str(kaisuu)
                vendor_data["henkou_flag"] = f"第{kaisuu}回変更"
            else:
                vendor_data["henkou_kaisuu"] = "0"
                vendor_data["henkou_flag"] = ""

            # ── 契約日（「当初」→「変更1」で上書き判定） ──
            if contract_date_row:
                raw_date = _cell_raw(ws, contract_date_row, base_col)
                date_dict = _parse_contract_date(raw_date)
                # 変更1行があれば上書き判定
                if contract_change_row:
                    change_raw = _cell_raw(ws, contract_change_row, base_col)
                    if change_raw is not None and str(change_raw).strip() != "":
                        date_dict = _parse_contract_date(change_raw)
                        logger.info("  契約日: 変更1の値を採用 (行%d)", contract_change_row)
                vendor_data.update(date_dict)
            else:
                vendor_data.update({
                    "contract_year": None,
                    "contract_month": None,
                    "contract_day": None,
                })

            # ── 工期（「当初」→「変更1」で上書き判定） ──
            if kouki_row:
                raw_kouki = _cell_raw(ws, kouki_row, base_col)
                kouki_dict = _parse_kouki(raw_kouki)
                # 変更1行があれば上書き判定
                if kouki_change_row:
                    change_raw = _cell_raw(ws, kouki_change_row, base_col)
                    if change_raw is not None and str(change_raw).strip() != "":
                        kouki_dict = _parse_kouki(change_raw)
                        logger.info("  工期: 変更1の値を採用 (行%d)", kouki_change_row)
                vendor_data.update(kouki_dict)
            else:
                vendor_data.update({
                    "kouki_start_year": None, "kouki_start_month": None,
                    "kouki_start_day": None,
                    "kouki_end_year": None, "kouki_end_month": None,
                    "kouki_end_day": None,
                })

            # ── 金額（「当初」取得 →「変更1」があれば差額計算） ──
            # 当初の金額を取得（スキャン範囲: 当初行～変更1行の手前）
            tosho_end = (kingaku_change_row - 1) if kingaku_change_row else max_row
            kingaku_tosho = _extract_kingaku_direct(
                ws, base_col, max_row,
                start_row=kingaku_tosho_row or 1,
                end_row=tosho_end,
            )

            if kingaku_change_row:
                # 変更1の金額を取得（スキャン範囲: 変更1行から5行分）
                kingaku_henkou = _extract_kingaku_direct(
                    ws, base_col, max_row,
                    start_row=kingaku_change_row,
                    end_row=kingaku_change_row + 5,
                )
                henkou_ukeoi = _safe_int(kingaku_henkou.get("kingaku_ukeoi"))

                if henkou_ukeoi > 0:
                    # 変更あり → 差額を計算
                    kingaku_found: dict[str, str | None] = {}
                    for k in ("kingaku_koji", "kingaku_zei", "kingaku_ukeoi"):
                        diff = _safe_int(kingaku_henkou[k]) - _safe_int(kingaku_tosho[k])
                        kingaku_found[k] = str(diff)
                    # 増減方向を判定（合計の差額で判定）
                    diff_ukeoi = _safe_int(kingaku_found["kingaku_ukeoi"])
                    if diff_ukeoi >= 0:
                        kingaku_found["kingaku_direction"] = "増"
                    else:
                        kingaku_found["kingaku_direction"] = "減"
                    logger.info(
                        "  金額: 差額を採用（変更1 - 当初）→ 工事価格=%s, 消費税=%s, 合計=%s [%s]",
                        kingaku_found["kingaku_koji"],
                        kingaku_found["kingaku_zei"],
                        kingaku_found["kingaku_ukeoi"],
                        kingaku_found["kingaku_direction"],
                    )
                else:
                    # 変更1の合計が空 or 0 → 当初をそのまま採用
                    kingaku_found = kingaku_tosho
                    kingaku_found["kingaku_direction"] = None
                    logger.info("  金額: 変更1の合計が空/0 → 当初の金額を採用")
            else:
                # 変更1行自体が存在しない → 当初をそのまま採用
                kingaku_found = kingaku_tosho
                kingaku_found["kingaku_direction"] = None

            vendor_data.update(kingaku_found)
            for k in ("kingaku_koji", "kingaku_zei", "kingaku_ukeoi"):
                if kingaku_found.get(k) is None:
                    logger.warning("  金額未取得: %s (業者='%s')", k, company)

            # ── 関連シート（排他割り当て済みの結果を使用） ──
            related = sheet_assignment.get(
                company, {"nairaku_sheet": None, "joken_sheet": None},
            )
            vendor_data.update(related)

            if related["nairaku_sheet"]:
                logger.info("  内訳書シート: '%s'", related["nairaku_sheet"])
            else:
                logger.warning("  内訳書シートが見つかりません: 業者='%s'", company)

            if related["joken_sheet"]:
                logger.info("  契約条件書シート: '%s'", related["joken_sheet"])
            else:
                logger.warning("  契約条件書シートが見つかりません: 業者='%s'", company)

            results.append(vendor_data)
            logger.info("業者 %d 抽出完了: %s", col_idx + 1, company)

            # ── 業者データ可視化（全フィールドをダンプ出力） ──
            logger.info("--- 業者 %d データダンプ開始: %s ---", col_idx + 1, company)
            for k, v in vendor_data.items():
                logger.info("  %-25s : %s", k, v)
            logger.info("--- 業者 %d データダンプ終了 ---", col_idx + 1)

        logger.info("Excel 読み込み完了: %d 社分のデータを抽出", len(results))
        return results

    finally:
        wb.close()
