"""
LibreOffice Headless PDF 変換モジュール

Session 0 環境における Excel COM (win32com) の構造的限界を回避するため、
LibreOffice の Headless モードを使用して Excel シートを PDF に変換する。

設計方針:
- win32com は一切使用しない（Session 0 安全性を 100% 保証）
- openpyxl で対象シートのみを含む一時ブックを作成
- subprocess で soffice --headless を呼び出し PDF 変換
- 一時ファイルは try-finally + tempfile で確実に削除

依存関係:
- openpyxl  : Excel ファイルの読み書き（純粋 Python・COM 不要）
- subprocess: LibreOffice 呼び出し
"""
from __future__ import annotations

import logging
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path

from . import config

logger = logging.getLogger(__name__)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  例外クラス
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class LibreOfficeNotFoundError(FileNotFoundError):
    """LibreOffice の実行ファイルが見つからない場合に送出される例外。"""

    def __init__(self, path: str):
        self.path = path
        super().__init__(
            f"LibreOffice が見つかりません: {path}\n"
            f".env の LIBREOFFICE_PATH を確認してください。"
        )


class LibreOfficeConversionError(RuntimeError):
    """LibreOffice での PDF 変換が失敗した場合に送出される例外。"""
    pass


class LibreOfficeTimeoutError(TimeoutError):
    """LibreOffice の PDF 変換がタイムアウトした場合に送出される例外。"""

    def __init__(self, timeout: int, sheet_name: str):
        self.timeout = timeout
        self.sheet_name = sheet_name
        super().__init__(
            f"LibreOffice PDF 変換がタイムアウトしました "
            f"（{timeout}秒, シート: '{sheet_name}'）"
        )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  内部ユーティリティ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _validate_libreoffice_path() -> Path:
    """
    LibreOffice の実行パスを検証して返す。

    Returns
    -------
    Path
        検証済みの soffice 実行パス。

    Raises
    ------
    LibreOfficeNotFoundError
        指定パスに soffice が存在しない場合。
    """
    lo_path = Path(config.LIBREOFFICE_PATH)

    # 絶対パスが指定されている場合はファイル存在チェック
    if lo_path.is_absolute():
        if not lo_path.exists():
            raise LibreOfficeNotFoundError(str(lo_path))
        return lo_path

    # 相対パスまたはコマンド名の場合は shutil.which で探索
    resolved = shutil.which(str(lo_path))
    if resolved is None:
        raise LibreOfficeNotFoundError(str(lo_path))
    return Path(resolved)


def _extract_sheet_to_temp_xlsx(
    excel_path: Path,
    sheet_name: str,
    temp_dir: Path,
) -> Path:
    """
    openpyxl で対象シートのみを含む一時 .xlsx ファイルを作成する。

    Parameters
    ----------
    excel_path : Path
        元の Excel ファイル。
    sheet_name : str
        抽出するシート名。
    temp_dir : Path
        一時ファイルの出力先ディレクトリ。

    Returns
    -------
    Path
        生成された一時 .xlsx ファイルのパス。

    Raises
    ------
    ValueError
        指定シートが元ファイルに存在しない場合。
    """
    import openpyxl

    logger.debug("openpyxl でシート抽出開始: '%s' from %s", sheet_name, excel_path.name)

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    try:
        # シート存在チェック
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"シート '{sheet_name}' が見つかりません。"
                f"利用可能なシート: [{available}]"
            )

        # 対象シート以外をすべて削除
        for name in wb.sheetnames:
            if name != sheet_name:
                del wb[name]

        # ASCII 安全なファイル名で一時ファイルに保存
        safe_name = f"lo_temp_{uuid.uuid4().hex[:8]}.xlsx"
        temp_xlsx = temp_dir / safe_name
        wb.save(str(temp_xlsx))
        logger.debug("一時ブック作成完了: %s (シート: '%s')", temp_xlsx.name, sheet_name)
        return temp_xlsx
    finally:
        wb.close()


def _run_libreoffice_convert(
    soffice_path: Path,
    input_xlsx: Path,
    output_dir: Path,
    timeout: int,
) -> Path:
    """
    LibreOffice Headless で .xlsx → .pdf 変換を実行する。

    Parameters
    ----------
    soffice_path : Path
        soffice 実行ファイルのパス。
    input_xlsx : Path
        入力 .xlsx ファイル。
    output_dir : Path
        PDF 出力先ディレクトリ。
    timeout : int
        タイムアウト（秒）。

    Returns
    -------
    Path
        生成された PDF ファイルのパス。

    Raises
    ------
    LibreOfficeConversionError
        変換プロセスがエラー終了した場合。
    LibreOfficeTimeoutError
        変換がタイムアウトした場合。
    """
    # LibreOffice は --outdir に PDF を出力する
    # 出力ファイル名は入力ファイルの拡張子を .pdf に置換したものになる
    cmd = [
        str(soffice_path),
        "--headless",
        "--norestore",
        "--nofirststartwizard",
        "--convert-to", "pdf",
        "--outdir", str(output_dir),
        str(input_xlsx),
    ]

    logger.info(
        "LibreOffice 変換実行: %s → %s",
        input_xlsx.name, output_dir,
    )
    logger.debug("コマンド: %s", " ".join(cmd))

    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            timeout=timeout,
            # Windows 環境では CREATE_NO_WINDOW でコンソールウィンドウの表示を防止
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except subprocess.TimeoutExpired as exc:
        logger.error(
            "LibreOffice タイムアウト (%d秒): %s", timeout, input_xlsx.name,
        )
        raise LibreOfficeTimeoutError(timeout, input_xlsx.stem) from exc

    # 標準出力・標準エラーをログに記録
    stdout_text = proc.stdout.decode("utf-8", errors="replace").strip()
    stderr_text = proc.stderr.decode("utf-8", errors="replace").strip()

    if stdout_text:
        logger.debug("LibreOffice stdout: %s", stdout_text)
    if stderr_text:
        logger.warning("LibreOffice stderr: %s", stderr_text)

    if proc.returncode != 0:
        raise LibreOfficeConversionError(
            f"LibreOffice がエラー終了しました (returncode={proc.returncode})。\n"
            f"stderr: {stderr_text}"
        )

    # 出力 PDF パスを推定（入力ファイル名の拡張子を .pdf に置換）
    expected_pdf = output_dir / input_xlsx.with_suffix(".pdf").name

    if not expected_pdf.exists():
        raise LibreOfficeConversionError(
            f"LibreOffice は正常終了しましたが、PDF が見つかりません: {expected_pdf}"
        )

    logger.info("LibreOffice PDF 生成完了: %s", expected_pdf.name)
    return expected_pdf


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  公開 API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def convert_excel_sheets_to_pdf(
    excel_path: Path,
    sheet_names: list[str],
    output_pdf_paths: list[Path],
    *,
    timeout: int | None = None,
) -> list[Path]:
    """
    Excel ファイルから指定シートを 1 枚ずつ PDF に変換する。
    （LibreOffice Headless 版 — Session 0 安全）

    office_com.convert_excel_sheets_to_pdf() と同じインターフェースを持ち、
    ドロップイン置換が可能。

    Parameters
    ----------
    excel_path : Path
        元の Excel ファイル（.xlsx / .xlsm）の絶対パス。
    sheet_names : list[str]
        PDF 化したいシート名のリスト。
    output_pdf_paths : list[Path]
        出力先 PDF の絶対パスのリスト（sheet_names と同順・同数）。
    timeout : int | None
        LibreOffice 変換タイムアウト（秒）。
        None の場合は config.LIBREOFFICE_TIMEOUT を使用。

    Returns
    -------
    list[Path]
        実際に生成できた PDF パスのリスト。

    Raises
    ------
    ValueError
        sheet_names と output_pdf_paths の数が一致しない場合。
    LibreOfficeNotFoundError
        LibreOffice が見つからない場合。
    """
    if len(sheet_names) != len(output_pdf_paths):
        raise ValueError(
            f"sheet_names ({len(sheet_names)}) と "
            f"output_pdf_paths ({len(output_pdf_paths)}) の数が一致しません"
        )

    if not sheet_names:
        return []

    # ── パス検証（全シート処理前に 1 回だけ実行） ──
    soffice_path = _validate_libreoffice_path()
    effective_timeout = timeout if timeout is not None else config.LIBREOFFICE_TIMEOUT

    logger.info(
        "LibreOffice PDF 変換開始: %s → %d シート",
        excel_path.name, len(sheet_names),
    )

    created: list[Path] = []

    # ── 一時ディレクトリ（自動削除保証） ──
    # tempfile.TemporaryDirectory は finally での rmtree を保証する
    temp_dir_path: Path | None = None
    temp_dir_obj = tempfile.TemporaryDirectory(
        prefix="lo_convert_",
        dir=str(config.COM_TEMP_DIR),
    )

    try:
        temp_dir_path = Path(temp_dir_obj.name)
        logger.debug("一時ディレクトリ作成: %s", temp_dir_path)

        for sheet_name, output_path in zip(sheet_names, output_pdf_paths):
            try:
                # 1. openpyxl で対象シートのみの一時ブック作成
                temp_xlsx = _extract_sheet_to_temp_xlsx(
                    excel_path, sheet_name, temp_dir_path,
                )

                # 2. LibreOffice で PDF 変換
                #    中間 PDF は一時ディレクトリに生成される
                intermediate_pdf = _run_libreoffice_convert(
                    soffice_path, temp_xlsx, temp_dir_path, effective_timeout,
                )

                # 3. 中間 PDF を最終出力先に移動
                output_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(intermediate_pdf), str(output_path))

                created.append(output_path)
                logger.info(
                    "シート PDF 変換完了: '%s' → %s",
                    sheet_name, output_path.name,
                )

            except ValueError as exc:
                # シートが見つからない場合 — スキップして次へ
                logger.warning("シート未検出 — スキップ: %s", exc)
                continue

            except (LibreOfficeConversionError, LibreOfficeTimeoutError) as exc:
                logger.error(
                    "LibreOffice 変換失敗 (シート '%s'): %s",
                    sheet_name, exc,
                )
                continue

            except Exception:
                logger.error(
                    "予期しないエラー (シート '%s')", sheet_name, exc_info=True,
                )
                continue

            finally:
                # 一時 xlsx は各シート処理後に即削除（ディスク節約）
                if "temp_xlsx" in locals() and temp_xlsx.exists():
                    try:
                        temp_xlsx.unlink()
                    except OSError:
                        logger.debug("一時 xlsx 削除失敗: %s", temp_xlsx)

    finally:
        # ── 一時ディレクトリ全体を確実に削除 ──
        try:
            temp_dir_obj.cleanup()
            logger.debug("一時ディレクトリ削除完了")
        except Exception:
            logger.warning(
                "一時ディレクトリ削除失敗: %s", temp_dir_path, exc_info=True,
            )

    logger.info(
        "LibreOffice PDF 変換完了: %d/%d シート成功",
        len(created), len(sheet_names),
    )

    return created
