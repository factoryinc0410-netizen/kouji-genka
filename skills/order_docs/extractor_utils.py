"""
extractor.py から切り出した低レベル共有ユーティリティ群。

このモジュールは下記2グループから成る:

  1) 純粋ユーティリティ — Worksheet/Workbook に依存せず、入力値のみで結果が決まる
     関数（_normalize, _clean_amount, _format_wareki, _parse_kouki ...）。

  2) Cell アクセスラッパー — Worksheet/Cell を引数に取るが、
     Worksheet API を try/except で薄く包むだけのリーフユーティリティ
     （_cell_str, _cell_raw, _cell_value_strip ...）。
     extractor.py 以外（例: vml_utils.py）からも共有したいので、
     循環 import を避けるためここに置く。

extractor.py からは re-export することで、既存の
`from skills.order_docs.extractor import _normalize` 等のインポートを
壊さずに済む。
"""
from __future__ import annotations

import logging
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  定数
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# 法人格除去パターン（シート名・業者名のあいまい比較で剥がす接尾辞）
_CORPORATE_SUFFIXES = [
    "株式会社", "有限会社", "合同会社", "合資会社", "合名会社",
    "(株)", "(有)", "(合)",
    "一般社団法人", "一般財団法人", "公益社団法人", "公益財団法人",
    "NPO法人", "特定非営利活動法人",
]

# Excel の日付シリアル値の起点: 1899-12-30
_EXCEL_EPOCH = datetime(1899, 12, 30)

# 西暦 → 令和への変換オフセット（西暦 - 2018 = 令和年）
_REIWA_OFFSET = 2018


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  文字列正規化
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _normalize(text: str) -> str:
    """全角→半角、空白・改行・ゼロ幅文字除去、括弧統一など揺れを吸収した比較用文字列を返す。"""
    s = unicodedata.normalize("NFKC", text)
    # 空白・改行・タブ・全角スペース・ノーブレークスペース・ゼロ幅文字を除去
    s = re.sub(r"[\s　 ​‌‍﻿\r\n\t]+", "", s)
    s = s.replace("（", "(").replace("）", ")")    # 全角丸括弧→半角
    s = s.replace("【", "[").replace("】", "]")    # 全角隅付き→半角角括弧
    return s


def _extract_core_name(company: str) -> str:
    """
    法人格・空白を除去した「コアな名称」を返す。
    例: "株式会社　法面" → "法面"
         "有限会社 山田建設" → "山田建設"
    """
    s = _normalize(company)
    for suffix in _CORPORATE_SUFFIXES:
        s = s.replace(_normalize(suffix), "")
    # 残った空白も除去
    s = re.sub(r"[\s　]+", "", s)
    return s


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Excel シリアル値 → datetime
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _is_excel_serial(value: Any) -> bool:
    """値が Excel の日付シリアル値と思われるかを判定する。"""
    if isinstance(value, (int, float)):
        # 1900年〜2100年の範囲: シリアル値 1 ～ 73415
        return 1 <= value <= 73415
    if isinstance(value, str):
        # 数字のみの文字列で5桁程度
        stripped = value.strip()
        if re.fullmatch(r"\d{4,6}", stripped):
            num = int(stripped)
            return 1 <= num <= 73415
    return False


def _serial_to_datetime(value: Any) -> datetime:
    """Excel シリアル値を Python datetime に変換する。"""
    num = int(float(str(value)))
    return _EXCEL_EPOCH + timedelta(days=num)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  金額・数値ユーティリティ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _clean_amount(raw: Any) -> str | None:
    """
    金額の生値をクリーニングして文字列で返す。
    - 数値型 (int/float) はそのまま文字列化
    - 文字列の場合はカンマ・円マーク・¥・スペースを除去して int 変換を試行
    - 変換できなければ None
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        # 小数点以下がない場合は整数文字列に
        if isinstance(raw, float) and raw == int(raw):
            return str(int(raw))
        return str(raw)
    # 文字列の場合: カンマ・円・¥・スペースを除去
    s = str(raw).strip()
    s = re.sub(r"[,，、円\xa5￥\s　]+", "", s)
    if not s:
        return None
    try:
        return str(int(s))
    except ValueError:
        try:
            return str(int(float(s)))
        except ValueError:
            logger.warning("金額変換失敗: '%s' (元値: %s)", s, repr(raw))
            return None


def _safe_int(s: str | None) -> int:
    """金額文字列を安全に int に変換する。None/空/変換不可は 0 を返す。"""
    if not s:
        return 0
    try:
        return int(re.sub(r"[^\d\-]", "", s))
    except ValueError:
        return 0


def _safe_float(value: Any) -> float | None:
    """セル値を float に変換する。None / 空 / 変換不可は None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("，", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Cell アクセスラッパー (Worksheet / Cell を薄く包むリーフ)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _cell_str(ws: Worksheet, row: int, col: int) -> str | None:
    """セルの値を文字列で返す。空や None は None。

    デフォルトでは前後空白を strip する（後方互換）。
    内訳書のように先頭インデント・中間空白を保持したい場合は
    `_cell_str_preserve()` を使用すること。
    """
    try:
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None
    except Exception:
        return None


def _cell_str_preserve(ws: Worksheet, row: int, col: int) -> str | None:
    """セルの値を空白を保持したまま文字列で返す。

    先頭の全角/半角空白（インデント）と中間空白（例: "土　工"）を
    そのまま維持する。完全に空（None or "" のみ）の場合のみ None。

    Excel 上で「見た目の余白」が意味を持つケース（内訳書の工種名等）で使う。
    """
    try:
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        s = str(val)
        # 完全空文字のみ None 扱い（rstrip せず "   " も保持する）
        return s if s != "" else None
    except Exception:
        return None


def _cell_raw(ws: Worksheet, row: int, col: int) -> Any:
    """セルの生値を返す（数値比較や日付変換用）。"""
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


# `_cell_str` / `_cell_str_preserve` は (ws, row, col) インタフェースだが、
# `ws.iter_rows(values_only=False)` から受け取る Cell を直接評価したい
# 場合に使うバリアント。ws.cell() のランダムアクセスを排除できる。

def _cell_value_preserve(cell: Any) -> str | None:
    """Cell オブジェクトから空白を保持した文字列を返す。

    `_cell_str_preserve` の Cell 版。先頭/中間空白を維持し、
    完全空（None or ""）のみ None を返す。
    iter_rows(values_only=False) の戻り値セルに対して使用する。
    """
    try:
        val = cell.value
    except Exception:
        return None
    if val is None:
        return None
    s = str(val)
    return s if s != "" else None


def _cell_value_strip(cell: Any) -> str | None:
    """Cell オブジェクトから前後空白を strip した文字列を返す。

    `_cell_str` の Cell 版。単位列など「空白が意味を持たない」列向け。
    """
    try:
        val = cell.value
    except Exception:
        return None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  和暦・日付パーサー
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _format_wareki(d: Any) -> str:
    """datetime/date を「令和X年Y月Z日」形式の和暦文字列に整形する。

    Parameters
    ----------
    d : datetime | date | None
        日付オブジェクト。None や非日付型はそのまま文字列化して返す（壊れにくさ優先）。

    Returns
    -------
    str
        例: datetime(2026, 1, 28) → '令和8年1月28日'
        令和元年 (2019) は '令和1年' として出力（慣例的な表記）。
        平成以前の日付は `_format_wareki` の守備範囲外なので西暦フォールバック。
    """
    if d is None:
        return ""
    if isinstance(d, datetime):
        y, mo, da = d.year, d.month, d.day
    elif hasattr(d, "year") and hasattr(d, "month") and hasattr(d, "day"):
        y, mo, da = d.year, d.month, d.day
    else:
        return str(d).strip()
    # 令和元年 (2019-05-01) 以降のみサポート
    if y >= 2019:
        reiwa = y - _REIWA_OFFSET
        return f"令和{reiwa}年{mo}月{da}日"
    # フォールバック: 西暦表示（データ異常への保険）
    return f"{y}年{mo}月{da}日"


def _parse_contract_date(raw: Any) -> dict[str, str | None]:
    """
    契約日セルの値をパースし、令和年・月・日に分割して返す。
    対応形式:
      - datetime オブジェクト（openpyxl が返す）
      - Excel シリアル値 (int/float, 例: 46091)
      - 文字列 "2025-12-01", "2025/12/1", "令和7年12月1日" 等
      - 数字だけの文字列 "46091"（シリアル値として変換）
    """
    result: dict[str, str | None] = {
        "contract_year": None,
        "contract_month": None,
        "contract_day": None,
    }

    if raw is None:
        return result

    # datetime 型
    if isinstance(raw, datetime):
        result["contract_year"] = str(raw.year - _REIWA_OFFSET)
        result["contract_month"] = str(raw.month)
        result["contract_day"] = str(raw.day)
        return result

    # Excel シリアル値（int / float）
    if _is_excel_serial(raw):
        dt = _serial_to_datetime(raw)
        result["contract_year"] = str(dt.year - _REIWA_OFFSET)
        result["contract_month"] = str(dt.month)
        result["contract_day"] = str(dt.day)
        logger.info("契約日シリアル値変換: %s → %s", raw, dt.strftime("%Y-%m-%d"))
        return result

    text = str(raw).strip()
    if not text:
        return result

    # 令和X年X月X日
    m = re.search(r"令和\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", text)
    if m:
        result["contract_year"] = m.group(1)
        result["contract_month"] = m.group(2)
        result["contract_day"] = m.group(3)
        return result

    # 西暦 YYYY-MM-DD or YYYY/MM/DD
    m = re.search(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", text)
    if m:
        year = int(m.group(1))
        result["contract_year"] = str(year - _REIWA_OFFSET)
        result["contract_month"] = m.group(2)
        result["contract_day"] = m.group(3)
        return result

    logger.warning("契約日パース失敗: '%s'", text)
    return result


def _parse_kouki(raw: Any) -> dict[str, str | None]:
    """
    工期セルの値をパースし、開始・終了の年月日を返す。
    対応形式:
      - "令和7年12月8日〜令和10年3月31日"
      - "令和7年12月8日～令和10年3月31日"（全角チルダ）
      - "2025/12/8〜2028/3/31"
      - datetime 型（単一日の場合は開始=終了とする）
    """
    result: dict[str, str | None] = {
        "kouki_start_year": None,
        "kouki_start_month": None,
        "kouki_start_day": None,
        "kouki_end_year": None,
        "kouki_end_month": None,
        "kouki_end_day": None,
    }

    if raw is None:
        return result

    if isinstance(raw, datetime):
        result["kouki_start_year"] = str(raw.year - _REIWA_OFFSET)
        result["kouki_start_month"] = str(raw.month)
        result["kouki_start_day"] = str(raw.day)
        return result

    # Excel シリアル値（単一日の場合）
    if _is_excel_serial(raw):
        dt = _serial_to_datetime(raw)
        result["kouki_start_year"] = str(dt.year - _REIWA_OFFSET)
        result["kouki_start_month"] = str(dt.month)
        result["kouki_start_day"] = str(dt.day)
        logger.info("工期シリアル値変換: %s → %s", raw, dt.strftime("%Y-%m-%d"))
        return result

    text = str(raw).strip()
    if not text:
        return result

    # 区切り文字（〜, ～, ~, -, ー）でスプリット
    parts = re.split(r"[〜～~\-ー]", text, maxsplit=1)

    def _extract_date(s: str, prefix: str) -> None:
        s = s.strip()
        # 令和パターン
        m = re.search(r"令和\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", s)
        if m:
            result[f"{prefix}_year"] = m.group(1)
            result[f"{prefix}_month"] = m.group(2)
            result[f"{prefix}_day"] = m.group(3)
            return
        # 西暦パターン
        m = re.search(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", s)
        if m:
            year = int(m.group(1))
            result[f"{prefix}_year"] = str(year - _REIWA_OFFSET)
            result[f"{prefix}_month"] = m.group(2)
            result[f"{prefix}_day"] = m.group(3)
            return
        logger.warning("工期日付パース失敗: '%s'", s)

    if len(parts) >= 1:
        _extract_date(parts[0], "kouki_start")
    if len(parts) >= 2:
        _extract_date(parts[1], "kouki_end")

    return result
