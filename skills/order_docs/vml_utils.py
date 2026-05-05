"""
契約条件書（joken）シートのテキスト・チェックボックス抽出ユーティリティ。

extractor.py から切り出した、VML（Vector Markup Language）解析と
契約条件書シート用のセル走査ヘルパを集約する。

含まれる主な機能:
  - VML チェックボックスの座標 → Excel 行番号変換
  - .xlsx 内 vmlDrawing*.vml をパースしてチェック状態を抽出
  - キーワードベースの行検索 / 隣接セル取得（汎用）
  - 契約条件書シートからスタンプ用テキストデータを一括抽出
    （extract_joken_text_data）

外部からは `from skills.order_docs.vml_utils import extract_joken_text_data`
のように直接 import する。
"""
from __future__ import annotations

import logging
import re
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .extractor_utils import _cell_str, _normalize

logger = logging.getLogger(__name__)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  VML 座標 → Excel 行番号マッピング定数
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# ── VML Y座標 → Excel 行番号(1-indexed) マッピング ──────────
# チェックボックスは VML 描画として配置されており、margin-top の
# pt 値から Excel 行番号を逆算する。
# 基準: VML y=71.0pt → Excel row 6, 間隔=13.0pt/行。
_VML_Y_BASE = 71.0
_VML_Y_STEP = 13.0
_VML_ROW_BASE = 6

# 元方列・下請列の VML X 座標 (margin-left の pt 値)
_VML_MOTOKATA_X = 316.5
_VML_SHITAUKE_X = 357.5
# 「9.適切な下請契約等」セクション用の特殊 X 座標
_VML_SECTION9_X = 357.5

# 費用負担チェックボックスが存在する Excel 行 (1-indexed)
_CHECKBOX_ROWS = [
    6, 7, 8,                            # 1.測量関係費
    10, 11, 12, 13, 14,                  # 2.安全関係費
    15, 16, 17, 18, 19, 20, 21, 22,      # 3.現場事務所
    24, 25, 26, 27, 28, 29, 30,          # 4.管理費用
    32, 33,                              # 5.環境改善
    35, 36, 37, 38, 39, 40, 41,          # 6.その他費用
    43, 44, 45,                          # 7.別途協議
    48, 49,                              # 8.その他
]


def _vml_y_to_row(vml_y: float) -> int:
    """VML の margin-top (pt) を Excel 行番号 (1-indexed) に変換する。"""
    return _VML_ROW_BASE + round((vml_y - _VML_Y_BASE) / _VML_Y_STEP)


def _extract_checkboxes_from_vml(
    excel_path: Path,
    sheet_name: str,
) -> dict[int, dict[str, bool]]:
    """
    契約条件書の VML からチェックボックスの状態を抽出する。

    .xlsx を ZIP として開き、vmlDrawing*.vml を解析して
    各行・各列のチェック状態を返す。
    openpyxl はフォームコントロールを読み取れないため、
    VML XML を直接パースする方式を採用する。

    Parameters
    ----------
    excel_path : Path
        Excel ファイルのパス。
    sheet_name : str
        対象シート名（VML ファイルの特定に使用）。

    Returns
    -------
    dict[int, dict[str, bool]]
        {Excel行番号: {"motokata": bool, "shitauke": bool}, ...}
    """
    result: dict[int, dict[str, bool]] = {}

    try:
        with zipfile.ZipFile(str(excel_path), "r") as zf:
            # VML ファイルを探す
            vml_files = [
                f for f in zf.namelist()
                if "vmlDrawing" in f and f.endswith(".vml")
            ]
            if not vml_files:
                logger.warning("VML ファイルが見つかりません: %s", excel_path.name)
                return result

            vml_content = zf.read(vml_files[0]).decode("utf-8")
    except Exception:
        logger.error("VML 読込失敗: %s", excel_path.name, exc_info=True)
        return result

    # 各 v:shape を解析
    shapes = vml_content.split("<v:shape ")[1:]

    for shape_text in shapes:
        # margin-left / margin-top を抽出
        ml_match = re.search(r"margin-left:\s*([\d.]+)pt", shape_text)
        mt_match = re.search(r"margin-top:\s*([\d.]+)pt", shape_text)
        if not ml_match or not mt_match:
            continue

        vml_x = float(ml_match.group(1))
        vml_y = float(mt_match.group(1))
        is_checked = "<x:Checked>1</x:Checked>" in shape_text

        # チェックされていないものは処理不要（デフォルトは全て未チェック扱い）
        if not is_checked:
            continue

        excel_row = _vml_y_to_row(vml_y)

        # 行エントリを初期化
        if excel_row not in result:
            result[excel_row] = {"motokata": False, "shitauke": False}

        # 列を判定
        if abs(vml_x - _VML_MOTOKATA_X) < 5.0:
            result[excel_row]["motokata"] = True
        elif abs(vml_x - _VML_SHITAUKE_X) < 5.0:
            result[excel_row]["shitauke"] = True

    logger.info(
        "VML チェックボックス抽出完了: %d 行にチェックあり",
        len(result),
    )
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Worksheet 内のキーワード検索 / 隣接セル取得
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _find_keyword_row(
    ws: Worksheet,
    keyword: str,
    *,
    scan_cols: tuple[int, ...] = (1, 2, 3, 4, 5),
    max_row: int | None = None,
) -> int | None:
    """
    ワークシート内でキーワードを含むセルの行番号を返す。

    Parameters
    ----------
    ws : Worksheet
        検索対象のワークシート。
    keyword : str
        検索キーワード（正規化後に部分一致で検索）。
    scan_cols : tuple[int, ...]
        検索対象の列番号 (1-indexed)。
    max_row : int | None
        検索範囲の最大行。None の場合は ws.max_row を使用。

    Returns
    -------
    int | None
        見つかった行番号 (1-indexed)、見つからなければ None。
    """
    limit = min(max_row or (ws.max_row or 100), 100)
    norm_kw = _normalize(keyword)

    for row in range(1, limit + 1):
        for col in scan_cols:
            val = _cell_str(ws, row, col)
            if val and norm_kw in _normalize(val):
                return row
    return None


def _read_adjacent_value(
    ws: Worksheet,
    row: int,
    label_col: int,
    *,
    scan_right: int = 4,
    scan_below: int = 1,
) -> str | None:
    """
    ラベルセルの右隣または下のセルから値を取得する。

    セル結合されている場合のフォールバック処理を含む。

    Parameters
    ----------
    row : int
        ラベルの行番号 (1-indexed)。
    label_col : int
        ラベルの列番号 (1-indexed)。
    scan_right : int
        右方向にスキャンする列数。
    scan_below : int
        下方向にスキャンする行数。
    """
    max_col = ws.max_column or 9

    # 右方向をスキャン
    for c in range(label_col + 1, min(label_col + scan_right + 1, max_col + 1)):
        v = _cell_str(ws, row, c)
        if v:
            return v

    # 下方向をスキャン（セル結合対応）
    for r in range(row + 1, row + scan_below + 1):
        v = _cell_str(ws, r, label_col)
        if v:
            return v
        # 下の行の右隣もチェック
        for c in range(label_col + 1, min(label_col + scan_right + 1, max_col + 1)):
            v = _cell_str(ws, r, c)
            if v:
                return v

    return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  契約条件書テキストデータ抽出（スタンプ方式 Route A 用）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def extract_joken_text_data(
    excel_path: Path,
    sheet_name: str,
) -> dict[str, str | None]:
    """
    契約条件書シートからスタンプ用の全テキストデータを抽出する。

    抽出対象:
    1. 現場代理人名
    2. 元方（左側）の商号又は名称・住所・氏名
    3. 下請（右側）の商号又は名称・住所・氏名
    4. 備考エリアのテキスト
    5. チェックボックスの状態（VML 解析）

    Parameters
    ----------
    excel_path : Path
        元の Excel ファイル。
    sheet_name : str
        契約条件書のシート名。

    Returns
    -------
    dict
        キー: スタンプフィールド名, 値: テキスト（見つからなければ None）。
        チェックボックスは "joken_checkboxes" キーに
        {行番号: {"motokata": bool, "shitauke": bool}} 形式で格納。
    """
    result: dict[str, Any] = {
        "joken_genba_dairinin": None,
        # 左側（元請）署名欄
        "joken_left_company": None,
        "joken_left_address": None,
        "joken_left_name": None,
        # 右側（下請）署名欄
        "joken_right_company": None,
        "joken_right_address": None,
        "joken_right_name": None,
        # 備考
        "joken_biko": None,
        # チェックボックス状態
        "joken_checkboxes": {},
    }

    # ── 1) チェックボックス抽出（ZIP/VML 直接解析） ──
    result["joken_checkboxes"] = _extract_checkboxes_from_vml(
        excel_path, sheet_name,
    )

    # ── 2) テキストデータ抽出（openpyxl） ──
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    except Exception:
        logger.error("契約条件書 Excel 読込失敗: %s", excel_path, exc_info=True)
        return result

    try:
        if sheet_name not in wb.sheetnames:
            logger.warning("契約条件書シート '%s' が存在しません", sheet_name)
            return result

        ws = wb[sheet_name]
        max_row = ws.max_row or 100
        max_col = ws.max_column or 9

        # ── 2a) 現場代理人 ──
        r = _find_keyword_row(ws, "現場代理人", max_row=10)
        if r is not None:
            for c in range(1, max_col + 1):
                v = _cell_str(ws, r, c)
                if v and "現場代理人" in _normalize(v):
                    val = _read_adjacent_value(ws, r, c)
                    if val:
                        result["joken_genba_dairinin"] = val
                        logger.debug("現場代理人: R%d → '%s'", r, val)
                    break

        # ── 2b) 署名欄（「商号又は名称」キーワードで検索） ──
        # 構造: 左側=A列ラベル→B列値、右側=D列ラベル→E/F列値
        # 「商号又は名称」の行を探し、その下に「住所」「氏名」がある
        shogo_rows: list[int] = []
        for row in range(max(max_row - 20, 1), max_row + 1):
            for col in (1, 2, 3, 4, 5, 6):
                v = _cell_str(ws, row, col)
                if v and "商号" in _normalize(v):
                    shogo_rows.append(row)
                    break

        # 最初の「商号又は名称」→ 左側（元請）
        # 2番目の「商号又は名称」→ 右側（下請）、同じ行の D列以降
        if shogo_rows:
            left_row = shogo_rows[0]
            # 左側: A列ラベル → B列に値（B列が結合でC列のこともある）
            result["joken_left_company"] = _read_adjacent_value(
                ws, left_row, 1, scan_right=2,
            )
            logger.debug("左側商号: R%d → '%s'", left_row, result["joken_left_company"])

            # 左側の下の行から住所・氏名を探す
            for dr in range(1, 6):
                r = left_row + dr
                v = _cell_str(ws, r, 1)
                if v is None:
                    continue
                norm = _normalize(v)
                if "住" in norm and "所" in norm and result["joken_left_address"] is None:
                    result["joken_left_address"] = _read_adjacent_value(
                        ws, r, 1, scan_right=2,
                    )
                    logger.debug("左側住所: R%d → '%s'", r, result["joken_left_address"])
                elif "氏" in norm and "名" in norm and result["joken_left_name"] is None:
                    result["joken_left_name"] = _read_adjacent_value(
                        ws, r, 1, scan_right=2,
                    )
                    logger.debug("左側氏名: R%d → '%s'", r, result["joken_left_name"])

            # 右側: 同じ行の D列以降
            for col in range(4, max_col + 1):
                v = _cell_str(ws, left_row, col)
                if v and "商号" in _normalize(v):
                    result["joken_right_company"] = _read_adjacent_value(
                        ws, left_row, col, scan_right=3,
                    )
                    logger.debug(
                        "右側商号: R%d:C%d → '%s'",
                        left_row, col, result["joken_right_company"],
                    )
                    # 右側の住所・氏名
                    for dr in range(1, 6):
                        r = left_row + dr
                        rv = _cell_str(ws, r, col)
                        if rv is None:
                            continue
                        norm = _normalize(rv)
                        if "住" in norm and "所" in norm and result["joken_right_address"] is None:
                            result["joken_right_address"] = _read_adjacent_value(
                                ws, r, col, scan_right=3,
                            )
                            logger.debug("右側住所: R%d → '%s'", r, result["joken_right_address"])
                        elif "氏" in norm and "名" in norm and result["joken_right_name"] is None:
                            result["joken_right_name"] = _read_adjacent_value(
                                ws, r, col, scan_right=3,
                            )
                            logger.debug("右側氏名: R%d → '%s'", r, result["joken_right_name"])
                    break

        # ── 2c) 備考エリア ──
        # 「備考」列 (G列=7) の全データを行ごとに収集
        biko_parts: list[str] = []
        biko_col: int | None = None

        # 「備考」ヘッダの列を検索（通常 G4=7列目）
        for col in range(1, max_col + 1):
            for row in range(1, 10):
                v = _cell_str(ws, row, col)
                if v and "備" in v and "考" in v:
                    biko_col = col
                    break
            if biko_col:
                break

        if biko_col:
            for row in range(6, max_row + 1):
                v = _cell_str(ws, row, biko_col)
                if v:
                    biko_parts.append(v.strip())

        if biko_parts:
            result["joken_biko"] = "\n".join(biko_parts)
            logger.debug("備考: %d 件取得", len(biko_parts))

    except Exception:
        logger.error("契約条件書データ抽出失敗: %s", sheet_name, exc_info=True)
    finally:
        wb.close()

    return result
