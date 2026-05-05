"""
契約条件書（terms / joken）データ抽出と契約変更回数スキャン。

extractor.py から切り出した、固定レイアウトに基づく契約条件書の
構造化抽出（TermsData）と、依頼書からの「変更回数」スキャンを集約する。

含まれる主な機能:
  - 契約条件書シートから TermsData を抽出（10 セクション × 各明細行）
  - チェックボックス値（✓/レ/v など）の判定
  - 注文書作成依頼書から業者ごとの契約変更回数を抽出

外部からは `from skills.order_docs.terms_extraction import extract_terms_data`
のように直接 import する。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from . import config
from .extractor_utils import (
    _banner,
    _cell_raw,
    _cell_str,
    _cell_str_preserve,
)
from .irai_scan_utils import _detect_vendor_base_cols
from .terms_models import TermsData, TermsItem, TermsParty, TermsSection

logger = logging.getLogger(__name__)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  契約変更回数の自動スキャン（注文書作成依頼書）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def scan_contract_change_count(
    excel_path: Path,
    *,
    sheet_keyword: str = "注文書作成依頼書",
    scan_rows_around: int = 15,
) -> dict[int, int]:
    """依頼書シートから業者ごとの契約変更回数を抽出する。

    構造（実物ファイル解析に基づく）:
      - 業者は横並び（業者1→C/D列, 業者2→E/F列, …, 業者5→K/L列）
      - "変更回数" ラベルの右隣セルに値（int）

    検出ロジック:
      1. 全シートから `sheet_keyword` を含むシートを特定
      2. 各セルを走査し "変更回数" 文字列を含むセルを発見したら、
         その右隣（+1列）セルの値を int に変換
      3. 左隣から業者番号を推定（業者基準列リストと照合）

    Returns
    -------
    dict[int, int]
        業者番号 (1-indexed) → 変更回数。値が空/0 の業者はエントリなし。
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel ファイルが見つかりません: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    try:
        # 対象シートを特定
        ws: Worksheet | None = None
        for name in wb.sheetnames:
            if sheet_keyword in name:
                ws = wb[name]
                break
        if ws is None:
            logger.warning("変更回数スキャン: シート '%s' 未検出", sheet_keyword)
            return {}

        # 業者基準列（1-indexed）: 自動検出 or config フォールバック
        max_row = ws.max_row or 50
        detected = _detect_vendor_base_cols(ws, max_row)
        base_cols = detected or config.EXCEL_MAP["vendor_base_cols"]

        # base_col → 業者番号
        col_to_vendor = {c: i + 1 for i, c in enumerate(base_cols)}

        result: dict[int, int] = {}
        max_col = ws.max_column or 20

        # 全セル走査（ヘッダ域付近のみ想定: row 1..max_row）
        for row in range(1, min(max_row + 1, 200)):
            for col in range(1, max_col + 1):
                val = _cell_str(ws, row, col)
                if not val or "変更回数" not in val:
                    continue
                # 右隣セルの値
                raw = _cell_raw(ws, row, col + 1)
                if raw is None or raw == "":
                    continue
                try:
                    count = int(float(str(raw)))
                except (ValueError, TypeError):
                    logger.debug("変更回数の数値変換失敗 row=%d col=%d raw=%r", row, col, raw)
                    continue
                if count <= 0:
                    continue
                # 業者番号の特定: col 自体 or 直近の base_col
                vendor_no = col_to_vendor.get(col)
                if vendor_no is None:
                    # ラベル列の1つ左が base_col であるケース
                    # 例: C38=ラベル, D38=値 → base_col=C(3), 業者1
                    vendor_no = col_to_vendor.get(col - 1)
                    # さらに許容: label 列自体が base_col から 1 ずれる場合
                if vendor_no is None:
                    # base_cols で最も近い列を採用
                    nearest = min(base_cols, key=lambda bc: abs(bc - col))
                    vendor_no = col_to_vendor[nearest]
                if vendor_no and vendor_no not in result:
                    result[vendor_no] = count

        logger.info("契約変更回数スキャン結果: %s", result)
        return result
    finally:
        wb.close()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  契約条件書データ抽出
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# セクション定義テーブル（実ファイル解析に基づく固定レイアウト）
# 各エントリ: (section_no, title_row, item_rows, layout)
_TERMS_SECTIONS_SPEC: list[tuple[int, int, list[int], str]] = [
    (1,  6,  [6, 7, 8],                            "std"),      # 1.測量関係費
    (2,  10, [10, 11, 12, 13, 14],                 "std"),      # 2.安全関係費
    (3,  15, [15, 16, 17, 18, 19, 20, 21, 22],     "std"),      # 3.現場事務所・仮設電力費
    (4,  24, [24, 25, 26, 27, 28, 29, 30],         "std"),      # 4.管理費用
    (5,  32, [32, 33],                             "std"),      # 5.現場環境改善費用
    (6,  35, [35, 36, 37, 38, 39, 40, 41],         "std"),      # 6.その他費用
    (7,  43, [43, 44, 45],                         "std"),      # 7.別途協議事項
    (8,  47, [48, 49],                             "std"),      # 8.その他
    (9,  50, [51, 52, 53, 54, 55, 56],             "wide"),     # 9.適切な下請契約等 (B:D=label, E:F=check)
    (10, 58, [58],                                 "single"),   # 10.4週8休等の実施
]


def _is_checked(cell_value: Any) -> bool:
    """セル値が ✓ チェック済みかを判定する。

    data validation dropdown で選択された "✓" 文字列を検出。
    空白・None はすべて未チェック扱い。
    """
    if cell_value is None:
        return False
    s = str(cell_value).strip()
    if not s:
        return False
    # ✓ (U+2713), ✔ (U+2714), レ点 (U+30EC), "v"/"V"/"x"/"X" 等を許容
    return s in ("✓", "✔", "レ", "v", "V", "×", "x", "X", "○", "●")


def extract_terms_data(
    excel_path: Path,
    vendor_index: int,
    *,
    sheet_keyword: str = "契約条件書",
) -> TermsData:
    """契約条件書シートから TermsData を抽出する。

    Parameters
    ----------
    excel_path : Path
        Excel ファイル。
    vendor_index : int
        業者番号 (1-indexed)。シート名 "契約条件書 （業者名N）" の N に対応。
    sheet_keyword : str
        シート名の識別キーワード。

    Returns
    -------
    TermsData
        抽出結果。該当シート未検出時は空の TermsData（sections=[]）。
    """
    result = TermsData()
    _banner("JOKEN", f"抽出開始 vendor_index={vendor_index} keyword='{sheet_keyword}'")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel ファイルが見つかりません: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    try:
        # シート検索: "契約条件書" + 業者番号 を含む
        target_name: str | None = None
        for name in wb.sheetnames:
            if sheet_keyword in name and str(vendor_index) in name:
                target_name = name
                break
        if target_name is None:
            # フォールバック: 業者番号が連番でシートに含まれていない場合、
            # "契約条件書" を含むシートを順次カウント
            joken_sheets = [n for n in wb.sheetnames if sheet_keyword in n]
            if 1 <= vendor_index <= len(joken_sheets):
                target_name = joken_sheets[vendor_index - 1]

        if target_name is None:
            _banner(
                "JOKEN-SKIPPED",
                f"契約条件書シート未検出 vendor_index={vendor_index} "
                f"keyword='{sheet_keyword}' sheets={wb.sheetnames}",
            )
            return result

        ws = wb[target_name]
        result.source_sheet = target_name

        # ── ヘッダ情報 ──
        # B2:D2 merged → 工事名値
        result.koji_kenmei = _cell_str_preserve(ws, 2, 2) or ""
        # G2 → 現場代理人氏名
        result.genba_dairinin = _cell_str_preserve(ws, 2, 7) or ""

        # ── セクション・明細行 ──
        for sec_no, title_row, item_rows, layout in _TERMS_SECTIONS_SPEC:
            # 大分類タイトル (A列)
            title = _cell_str(ws, title_row, 1) or ""
            section = TermsSection(number=sec_no, title=title, layout=layout)

            for r in item_rows:
                if layout == "std":
                    # C:D merged = label, E = 元方, F = 下請, G = 備考
                    label = _cell_str_preserve(ws, r, 3) or ""
                    mk = _is_checked(_cell_raw(ws, r, 5))
                    st = _is_checked(_cell_raw(ws, r, 6))
                    biko = _cell_str_preserve(ws, r, 7) or ""
                elif layout == "wide":
                    # B:D merged = label, E:F merged = check, G = 備考
                    label = _cell_str_preserve(ws, r, 2) or ""
                    mk = _is_checked(_cell_raw(ws, r, 5))
                    st = _is_checked(_cell_raw(ws, r, 6))
                    biko = _cell_str_preserve(ws, r, 7) or ""
                elif layout == "single":
                    # セクション10 (row 58): ラベルは A列, E/F でチェック, G で備考
                    label = ""  # セクションタイトル自体が明細を兼ねる
                    mk = _is_checked(_cell_raw(ws, r, 5))
                    st = _is_checked(_cell_raw(ws, r, 6))
                    biko = _cell_str_preserve(ws, r, 7) or ""
                else:
                    continue

                section.items.append(TermsItem(
                    excel_row=r,
                    label=label,
                    motokata_checked=mk,
                    shitauke_checked=st,
                    biko=biko,
                    layout=layout,
                ))

            result.sections.append(section)

        # ── サインブロック ──
        # B61:D61 → 共同企業体名
        result.motouke_group_name = _cell_str_preserve(ws, 61, 2) or ""
        # A62 → （代表構成員）
        koseiin = _cell_str_preserve(ws, 62, 1) or "（代表構成員）"
        result.daikyo_koseiin_label = koseiin

        # 左側（元請）: 行 62:住, 63:商, 64:氏 — C列に値
        result.motouke = TermsParty(
            address=_cell_str_preserve(ws, 62, 3) or "",
            company=_cell_str_preserve(ws, 63, 3) or "",
            name=_cell_str_preserve(ws, 64, 3) or "",
        )
        # 右側（下請）: 行 62:商, 63:住, 64:氏 — F列に値（F:G merged）
        result.shitauke = TermsParty(
            address=_cell_str_preserve(ws, 63, 6) or "",
            company=_cell_str_preserve(ws, 62, 6) or "",
            name=_cell_str_preserve(ws, 64, 6) or "",
        )

        logger.info(
            "契約条件書データ抽出完了: %s / vendor=%d / sections=%d, "
            "checked(元方)=%d, checked(下請)=%d",
            target_name, vendor_index, len(result.sections),
            sum(1 for s in result.sections for it in s.items if it.motokata_checked),
            sum(1 for s in result.sections for it in s.items if it.shitauke_checked),
        )
        if result.sections:
            total_items = sum(len(s.items) for s in result.sections)
            _banner(
                "JOKEN-SUCCESS",
                f"sheet='{target_name}' vendor_index={vendor_index} "
                f"sections={len(result.sections)} items={total_items}",
            )
        else:
            _banner(
                "JOKEN-FAILED",
                f"sheet='{target_name}' vendor_index={vendor_index} "
                f"sections=0 — Excel の _TERMS_SECTIONS_SPEC 対応行を確認してください",
            )

    except Exception as exc:
        _banner(
            "JOKEN-ERROR",
            f"vendor_index={vendor_index} 例外: {type(exc).__name__}: {exc}",
        )
        logger.error("契約条件書データ抽出失敗: %s / vendor=%d",
                     excel_path, vendor_index, exc_info=True)
    finally:
        wb.close()

    return result
