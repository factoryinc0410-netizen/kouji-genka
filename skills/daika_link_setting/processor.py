"""代価表リンク設定 — openpyxl で内部ハイパーリンクを付与する処理。

入力 Excel を読み込み、各シートの A〜D 列から「単/施/P/明 N 号」のターゲット
位置をスキャンしてマッピングを構築し、ソース側（本工事費表の基準列など）の
該当セルに同一ワークブック内ハイパーリンクを設定して保存する。

マッチングは全て NFKC 正規化 + 全空白除去 + 大文字統一を経た文字列で行う:
  - 「単 1 号」「単１号」「単1号」  → "単1号"
  - 「-0001 Ｐ 1 号」              → "-0001P1号"（抽出キー: "P1号"）
  - 「-0001 施 1 号」              → "-0001施1号"（抽出キー: "施1号"）
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

# 「単/施/P/明 + 数字 + 号」を抽出。NFKC 正規化＋空白除去済みの文字列に適用する想定。
_KEYWORD_RE = re.compile(r"(単|施|P|明)\d+号")

# シート上で「基準」ヘッダを探す走査範囲。
_HEADER_SCAN_MAX_ROW = 50
_HEADER_SCAN_MAX_COL = 30
_BASIS_HEADER_NORMALIZED = "基準"  # NFKC + 空白除去後の比較値

# ターゲット側スキャン範囲（A〜D 列）— ここに置かれた「単1号」「施1号」等が
# リンクの飛び先になる。「リンクの起点」を探す走査範囲とは厳格に分離する。
_TARGET_MIN_COL = 1
_TARGET_MAX_COL = 4

# リンクの起点（ソース）を走査する最小列番号 = E 列。
# ターゲット側の A〜D 列を除外することで、ターゲットセル自身が自分自身へ
# 自己リンクするのを防ぐ。「基準」列がシートによって M〜R など揺れても、
# E 列以降を一律スキャンすることで列ズレに堅牢にする（ヘッダ位置に依存しない）。
_SOURCE_MIN_COL = 5

# 未マッチキーワードの収集上限（UI で先頭 N 件を表示できれば十分）
_UNMATCHED_CAPACITY = 200


@dataclass
class LinkStats:
    """処理結果のサマリー。UI へそのまま渡す。"""

    sheets_scanned: int = 0
    targets_indexed: int = 0
    links_created: int = 0
    keywords_unmatched: list[str] = field(default_factory=list)
    basis_columns: dict[str, str] = field(default_factory=dict)
    duplicate_targets: list[str] = field(default_factory=list)


def normalize_key(value: object) -> str:
    """NFKC 正規化 → 全空白除去。Ｐ → P もこの過程で起こる。"""
    if value is None:
        return ""
    s = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", "", s)


def extract_keyword(normalized: str) -> str | None:
    """正規化済み文字列の中から最初に出現する `(単|施|P|明)\\d+号` を返す。"""
    if not normalized:
        return None
    m = _KEYWORD_RE.search(normalized)
    return m.group(0) if m else None


def _find_basis_column(ws: Worksheet) -> int | None:
    """正規化後に「基準」となるヘッダセルが現れる列番号を返す。

    シート先頭 50 行 × 30 列のみ走査。複数候補があれば最初のもの。
    """
    max_row = min(_HEADER_SCAN_MAX_ROW, ws.max_row or 0)
    max_col = min(_HEADER_SCAN_MAX_COL, ws.max_column or 0)
    if max_row == 0 or max_col == 0:
        return None
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col):
        for cell in row:
            if normalize_key(cell.value) == _BASIS_HEADER_NORMALIZED:
                return cell.column
    return None


def _resolve_anchor(ws: Worksheet, cell: Cell) -> str:
    """セルが結合範囲内なら結合範囲の左上セル参照を返す。"""
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= cell.row <= merged.max_row
                and merged.min_col <= cell.column <= merged.max_col):
            return ws.cell(merged.min_row, merged.min_col).coordinate
    return cell.coordinate


def _build_target_index(
    wb: Workbook, stats: LinkStats,
) -> dict[str, tuple[str, str]]:
    """A〜D 列をスキャンし、抽出キーワード → (シート名, アンカーセル参照) の辞書を返す。

    「ターゲット」とは、セル値そのものがキーワード（正規化後 == 抽出キー）と
    完全に一致する明細行を指す。プレフィックス付き（"-0001施1号" 等）は
    ソース側にしか現れない想定なのでインデックスからは除外する。
    """
    index: dict[str, tuple[str, str]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row or 0,
            min_col=_TARGET_MIN_COL, max_col=_TARGET_MAX_COL,
        ):
            for cell in row:
                if cell.value is None:
                    continue
                normalized = normalize_key(cell.value)
                kw = extract_keyword(normalized)
                if not kw or normalized != kw:
                    continue
                anchor = _resolve_anchor(ws, cell)
                if kw in index:
                    # 二重定義は最初に見つけた位置を採用しつつ統計に記録
                    if kw not in stats.duplicate_targets:
                        stats.duplicate_targets.append(kw)
                    continue
                index[kw] = (ws.title, anchor)
    return index


def _make_link_font(base: Font | None) -> Font:
    """既存フォントの属性（サイズ等）を保ちつつ、色と下線だけハイパーリンク風にする。"""
    if base is None:
        return Font(color="0000FF", underline="single")
    return Font(
        name=base.name,
        size=base.size,
        bold=base.bold,
        italic=base.italic,
        family=base.family,
        scheme=base.scheme,
        color="0000FF",
        underline="single",
    )


def _quote_sheet(name: str) -> str:
    """シート名に含まれるシングルクォートをエスケープした上で `'...'` で囲む。"""
    return "'" + name.replace("'", "''") + "'"


def _apply_links(
    wb: Workbook,
    target_index: dict[str, tuple[str, str]],
    stats: LinkStats,
) -> None:
    """各シートで E 列以降の全セルを走査し、キーワードに該当するものに
    ハイパーリンクを付与する。

    ヘッダ「基準」の列が M/N/O/P/Q/R 等にシートごとにズレても、ヘッダ列と
    実際のデータ列が異なっていても、E 列以降を一律スキャンすることで取りこぼし
    なくリンクが付く。A〜D 列はターゲット側なので走査対象外（_SOURCE_MIN_COL=E）。

    `_find_basis_column` で見つかったヘッダ列は stats.basis_columns に
    「参考情報」として記録するが、走査範囲の判定には使わない。
    """
    for ws in wb.worksheets:
        stats.sheets_scanned += 1
        # 報告用: ヘッダで見つかった「基準」列の文字（ユーザー UI で目視確認用）。
        # 見つからなくても処理は継続する。
        basis_col = _find_basis_column(ws)
        if basis_col is not None:
            stats.basis_columns[ws.title] = ws.cell(1, basis_col).column_letter

        max_col = ws.max_column or 0
        if max_col < _SOURCE_MIN_COL:
            continue  # シートが A〜D 列しか持たない → ソースなし

        max_row = ws.max_row or 0
        for row_idx in range(1, max_row + 1):
            for col_idx in range(_SOURCE_MIN_COL, max_col + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is None:
                    continue
                normalized = normalize_key(cell.value)
                kw = extract_keyword(normalized)
                if not kw:
                    continue
                target = target_index.get(kw)
                if target is None:
                    if (kw not in stats.keywords_unmatched
                            and len(stats.keywords_unmatched) < _UNMATCHED_CAPACITY):
                        stats.keywords_unmatched.append(kw)
                    continue
                target_sheet, target_anchor = target
                if target_sheet == ws.title and target_anchor == cell.coordinate:
                    continue  # 自己参照
                location = f"{_quote_sheet(target_sheet)}!{target_anchor}"
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    location=location,
                    display=str(cell.value),
                )
                cell.font = _make_link_font(cell.font)
                stats.links_created += 1


def process(
    source: BinaryIO | bytes | bytearray | str | Path,
    out_path: str | Path,
) -> LinkStats:
    """エントリポイント: source を読み込み、リンク付与後 out_path に保存する。

    Args:
        source: アップロードされた Excel のバイナリ / ファイルパス
        out_path: 出力先 .xlsx パス
    Returns:
        LinkStats: 処理サマリー
    """
    if isinstance(source, (bytes, bytearray)):
        buf: BinaryIO | str | Path = BytesIO(bytes(source))
    else:
        buf = source

    wb = load_workbook(buf, data_only=False, keep_vba=False)
    try:
        stats = LinkStats()
        target_index = _build_target_index(wb, stats)
        stats.targets_indexed = len(target_index)
        _apply_links(wb, target_index, stats)
        wb.save(str(out_path))
    finally:
        wb.close()
    return stats
