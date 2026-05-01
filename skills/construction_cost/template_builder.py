"""
工事日報集計スキル — 入力用日報テンプレート自動生成

DBのマスタデータからプルダウン付きExcelテンプレートを生成する。
  - 20日締め対応: シート順は「21日」〜「31日」「1日」〜「20日」
  - 「リスト一覧」シート（先頭・表示）: プルダウンソースの格納先
  - 列順: [作業員名, 現場名, 開始時間, 終了時間, 休憩時間, 潜水作業, 船舶作業, 備考欄]
  - プルダウン: 作業員名(A列), 現場名(B列), 開始/終了時間(C/D列), 休憩時間(E列), 潜水/船舶(F/G列)
"""
from __future__ import annotations

import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("skills.construction_cost.template_builder")

# ── フォント・スタイル定義 ──────────────────────────
FONT_NAME = "MS ゴシック"

_HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CELL_FONT = Font(name=FONT_NAME, size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
_INFO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_INFO_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")

# ── レイアウト定数 ────────────────────────────────
HEADER_ROW = 1
DATA_START = 2
DATA_ROWS = 30
ROW_HEIGHT = 20

# 列定義 (A〜H)
COLUMNS = {
    "作業員名": {"col": 1, "width": 16},
    "現場名":   {"col": 2, "width": 42},
    "開始時間": {"col": 3, "width": 10},
    "終了時間": {"col": 4, "width": 10},
    "休憩時間": {"col": 5, "width": 10},
    "潜水作業": {"col": 6, "width": 10},
    "船舶作業": {"col": 7, "width": 10},
    "備考欄":   {"col": 8, "width": 24},
}

# リスト一覧シートの列割当て
MASTER_SHEET_NAME  = "リスト一覧"
MASTER_COL_SITES   = 1  # A列: 現場名
MASTER_COL_WORKERS = 2  # B列: 作業員名
MASTER_COL_TIMES   = 3  # C列: 時刻リスト (05:00〜22:00 / 15分刻み)
MASTER_COL_BREAKS  = 4  # D列: 休憩リスト (00:00〜03:00 / 15分刻み)

# 20日締め: シート順
CLOSING_DAY = 20  # 締め日


def _generate_time_list() -> list[str]:
    """05:00 から 22:00 まで15分刻みの時刻文字列リストを生成する。"""
    times = []
    for h in range(5, 23):
        for m in (0, 15, 30, 45):
            if h == 22 and m > 0:
                break
            times.append(f"{h:02d}:{m:02d}")
    return times


def _generate_break_list() -> list[str]:
    """00:00 から 03:00 まで15分刻みの休憩時間リストを生成する。"""
    breaks = []
    for h in range(4):
        for m in (0, 15, 30, 45):
            if h == 3 and m > 0:
                break
            breaks.append(f"{h:02d}:{m:02d}")
    return breaks


def _apply_header(ws, row: int):
    """ヘッダ行にスタイルを適用する。"""
    for name, info in COLUMNS.items():
        cell = ws.cell(row=row, column=info["col"], value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _apply_data_borders(ws, start_row: int, end_row: int, col_count: int):
    """データ領域に罫線・フォント・行高さを適用する。"""
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN_BORDER
            cell.font = _CELL_FONT


def _set_column_widths(ws):
    """列幅を設定する。"""
    for info in COLUMNS.values():
        ws.column_dimensions[get_column_letter(info["col"])].width = info["width"]


# 動的追記対応: プルダウン参照範囲の最大行数
# 現場名・作業員名はユーザーがリスト一覧シートに追記できるよう余裕を持たせる
DROPDOWN_MAX_ROW = 500


def _make_range_formula(col_idx: int, count: int, *, dynamic: bool = False) -> str:
    """リスト一覧シートの列範囲を参照する数式を返す。

    Args:
        col_idx: 列番号 (1-indexed)
        count:   実データ件数
        dynamic: True の場合、追記に対応するため DROPDOWN_MAX_ROW まで参照する
    """
    col_letter = get_column_letter(col_idx)
    if dynamic:
        end_row = DROPDOWN_MAX_ROW
    elif count == 0:
        end_row = 2
    else:
        end_row = count + 1
    return f"'{MASTER_SHEET_NAME}'!${col_letter}$2:${col_letter}${end_row}"


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int, formula: str,
                  prompt: str = "プルダウンから選択してください"):
    """指定列のセル範囲にプルダウン（入力規則）を設定する。"""
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください。",
        showInputMessage=True,
        promptTitle="選択",
        prompt=prompt,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def _write_master_sheet(
    wb: Workbook,
    site_names: list[str],
    worker_names: list[str],
) -> tuple[int, int]:
    """リスト一覧シートにプルダウンソースを書き込む。

    Returns:
        (時刻リスト件数, 休憩リスト件数)
    """
    ws = wb.create_sheet(MASTER_SHEET_NAME)

    # A列: 現場名
    ws.cell(row=1, column=MASTER_COL_SITES, value="現場名")
    ws.cell(row=1, column=MASTER_COL_SITES).font = Font(name=FONT_NAME, bold=True)
    for i, name in enumerate(site_names, 2):
        ws.cell(row=i, column=MASTER_COL_SITES, value=name)

    # B列: 作業員名
    ws.cell(row=1, column=MASTER_COL_WORKERS, value="作業員名")
    ws.cell(row=1, column=MASTER_COL_WORKERS).font = Font(name=FONT_NAME, bold=True)
    for i, name in enumerate(worker_names, 2):
        ws.cell(row=i, column=MASTER_COL_WORKERS, value=name)

    # C列: 時刻リスト (05:00〜22:00 / 15分刻み)
    time_list = _generate_time_list()
    ws.cell(row=1, column=MASTER_COL_TIMES, value="時刻")
    ws.cell(row=1, column=MASTER_COL_TIMES).font = Font(name=FONT_NAME, bold=True)
    for i, t in enumerate(time_list, 2):
        ws.cell(row=i, column=MASTER_COL_TIMES, value=t)

    # D列: 休憩リスト (00:00〜03:00 / 15分刻み)
    break_list = _generate_break_list()
    ws.cell(row=1, column=MASTER_COL_BREAKS, value="休憩")
    ws.cell(row=1, column=MASTER_COL_BREAKS).font = Font(name=FONT_NAME, bold=True)
    for i, b in enumerate(break_list, 2):
        ws.cell(row=i, column=MASTER_COL_BREAKS, value=b)

    # 列幅を見やすく設定
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    return len(time_list), len(break_list)


def build_template(
    site_names: list[str],
    worker_names: list[str],
    target_month: str = "",
) -> BytesIO:
    """プルダウン付きの日報テンプレートExcelを生成して BytesIO で返す。

    Args:
        site_names:   現場名リスト（DBから取得済み）
        worker_names: 作業員名リスト（DBから取得済み）
        target_month: 対象年月（タイトル表示用）
    """
    wb = Workbook()
    wb.remove(wb.active)

    logger.info("テンプレート生成開始: 現場%d件, 作業員%d名", len(site_names), len(worker_names))

    # ── 1. MasterData シート ──────────────────────────
    time_count, break_count = _write_master_sheet(wb, site_names, worker_names)

    # プルダウン参照数式（現場名・作業員名は追記対応で広めの範囲）
    site_formula   = _make_range_formula(MASTER_COL_SITES,   len(site_names), dynamic=True)
    worker_formula = _make_range_formula(MASTER_COL_WORKERS, len(worker_names), dynamic=True)
    time_formula   = _make_range_formula(MASTER_COL_TIMES,   time_count)
    break_formula  = _make_range_formula(MASTER_COL_BREAKS,  break_count)

    # ── 2. 日別シート（20日締め: 21日〜31日, 1日〜20日）──
    total_cols = len(COLUMNS)
    day_order = list(range(CLOSING_DAY + 1, 32)) + list(range(1, CLOSING_DAY + 1))

    for day in day_order:
        sheet_name = f"{day}日"
        ws = wb.create_sheet(sheet_name)

        # ヘッダ（1行目）
        _apply_header(ws, HEADER_ROW)

        # データ領域の罫線 + 行高さ
        data_end = DATA_START + DATA_ROWS - 1
        _apply_data_borders(ws, DATA_START, data_end, total_cols)

        # 列幅
        _set_column_widths(ws)

        # ★ プルダウン設定
        _add_dropdown(ws, "A", DATA_START, data_end, worker_formula,
                      prompt="作業員名を選択")
        _add_dropdown(ws, "B", DATA_START, data_end, site_formula,
                      prompt="現場名を選択")
        _add_dropdown(ws, "C", DATA_START, data_end, time_formula,
                      prompt="開始時間を選択")
        _add_dropdown(ws, "D", DATA_START, data_end, time_formula,
                      prompt="終了時間を選択")
        _add_dropdown(ws, "E", DATA_START, data_end, break_formula,
                      prompt="休憩時間を選択")
        # 潜水作業・船舶作業: ●のドロップダウン（直接リスト指定）
        _add_dropdown(ws, "F", DATA_START, data_end, '"●"',
                      prompt="該当する場合は●を選択")
        _add_dropdown(ws, "G", DATA_START, data_end, '"●"',
                      prompt="該当する場合は●を選択")
        # 潜水作業・船舶作業列を中央揃え
        for r in range(DATA_START, data_end + 1):
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")

    # ── 3. シートの並び順を整理（リスト一覧を先頭に）──
    wb.move_sheet(MASTER_SHEET_NAME, offset=-(len(wb.sheetnames) - 1))
    # アクティブシートは最初の日別シート（21日）
    wb.active = wb.sheetnames.index(f"{CLOSING_DAY + 1}日")

    # ── 4. BytesIO に書き出し ─────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info("テンプレート生成完了: %d シート", len(wb.sheetnames))
    return buf
