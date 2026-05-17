"""
工事日報集計スキル — Excel出力

出力ファイル:
  1. 現場別原価管理表_YYYY-MM.xlsx
  2. 個人別集計表_YYYY-MM.xlsx
  3. ダッシュボードExcelエクスポート（StreamingResponse用）
"""
from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from skills.construction_cost.aggregator import AggregationResult

logger = logging.getLogger("skills.construction_cost.writer")

# スタイル定義
FONT_NAME = "MS ゴシック"
_HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
_CELL_FONT = Font(name=FONT_NAME, size=10)
_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
_BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
_MONEY_FORMAT = '#,##0'
_HOUR_FORMAT = '#,##0.0'
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WARNING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WARNING_FONT = Font(name=FONT_NAME, size=10, color="9C0006")


def _apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _apply_cell_style(cell, fmt=None):
    cell.font = _CELL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if fmt:
        cell.number_format = fmt


def write_site_cost_report(result: AggregationResult, output_dir: Path) -> Path:
    """現場別原価管理表を出力する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "現場別原価管理表"

    # タイトル行
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"現場別原価管理表（{result.target_month}）"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # ヘッダ（経費列を削除: 9列→8列）
    headers = [
        "現場名", "予算金額", "前月末累計", "人件費",
        "当月支払計", "累計金額", "予算残", "消化率(%)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    # データ行
    row = 4
    total_budget = 0
    total_cum_before = 0
    total_labor = 0
    total_monthly = 0
    total_cum_after = 0

    for sc in result.site_costs:
        ws.cell(row=row, column=1, value=sc.site_name)
        _apply_cell_style(ws.cell(row=row, column=1))

        for col, val in [
            (2, sc.budget), (3, sc.cumulative_before), (4, sc.labor_cost),
            (5, sc.monthly_cost), (6, sc.cumulative_after), (7, sc.remaining),
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            _apply_cell_style(cell, _MONEY_FORMAT)

        # 消化率
        rate = (sc.cumulative_after / sc.budget * 100) if sc.budget > 0 else 0
        cell = ws.cell(row=row, column=8, value=round(rate, 1))
        _apply_cell_style(cell, '0.0')

        # 予算超過の警告
        if sc.remaining < 0:
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = _WARNING_FILL
                ws.cell(row=row, column=c).font = _WARNING_FONT

        total_budget += sc.budget
        total_cum_before += sc.cumulative_before
        total_labor += sc.labor_cost
        total_monthly += sc.monthly_cost
        total_cum_after += sc.cumulative_after
        row += 1

    # 合計行
    ws.cell(row=row, column=1, value="合計")
    _apply_cell_style(ws.cell(row=row, column=1))
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for col, val in [
        (2, total_budget), (3, total_cum_before), (4, total_labor),
        (5, total_monthly), (6, total_cum_after),
        (7, total_budget - total_cum_after),
    ]:
        cell = ws.cell(row=row, column=col, value=val)
        _apply_cell_style(cell, _MONEY_FORMAT)
        cell.font = _BOLD_FONT

    # 列幅調整
    col_widths = [20, 15, 15, 15, 15, 15, 15, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output_path = output_dir / f"現場別原価管理表_{result.target_month}.xlsx"
    wb.save(str(output_path))
    logger.info("現場別原価管理表を出力: %s", output_path)
    return output_path


_GROUP_SUBTOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_GROUP_SUBTOTAL_FONT = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")


def write_worker_summary(result: AggregationResult, output_dir: Path) -> Path:
    """個人別集計表を出力する（グループ別小計付き）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "個人別集計表"

    # タイトル
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"個人別集計表（{result.target_month}）"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # ヘッダ（グループ列追加）
    headers = [
        "氏名", "グループ", "職種", "出勤日数", "基本(h)", "残業(h)",
        "合計時間(h)", "人件費",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    # データ行（グループ別小計付き）
    row = 4
    current_group = None
    group_cost = 0.0

    summaries = result.worker_summaries

    for idx, ws_row in enumerate(summaries):
        grp = ws_row.group_name or "未分類"

        # グループが変わったら小計行を挿入
        if current_group is not None and grp != current_group:
            row = _write_group_subtotal(ws, row, current_group, group_cost, len(headers))
            group_cost = 0.0

        current_group = grp
        total_hours = ws_row.total_basic + ws_row.total_overtime

        ws.cell(row=row, column=1, value=ws_row.worker_name)
        _apply_cell_style(ws.cell(row=row, column=1))
        ws.cell(row=row, column=2, value=grp)
        _apply_cell_style(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=ws_row.role)
        _apply_cell_style(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value=ws_row.days_worked)
        _apply_cell_style(ws.cell(row=row, column=4))

        for col, val in [
            (5, ws_row.total_basic), (6, ws_row.total_overtime),
            (7, total_hours),
        ]:
            cell = ws.cell(row=row, column=col, value=round(val, 2))
            _apply_cell_style(cell, _HOUR_FORMAT)

        cell = ws.cell(row=row, column=8, value=ws_row.labor_cost)
        _apply_cell_style(cell, _MONEY_FORMAT)

        group_cost += ws_row.labor_cost
        row += 1

    # 最後のグループの小計
    if current_group is not None:
        row = _write_group_subtotal(ws, row, current_group, group_cost, len(headers))

    # --- 現場別内訳シート ---
    ws2 = wb.create_sheet("現場別内訳")
    ws2.merge_cells("A1:F1")
    ws2["A1"].value = f"個人別 現場別内訳（{result.target_month}）"
    ws2["A1"].font = _TITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center")

    detail_headers = ["氏名", "グループ", "現場名", "基本(h)", "残業(h)", "合計(h)"]
    for col, h in enumerate(detail_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    _apply_header_style(ws2, 3, len(detail_headers))

    row = 4
    for ws_row in summaries:
        grp = ws_row.group_name or "未分類"
        for site_name, hours in sorted(ws_row.site_hours.items()):
            ws2.cell(row=row, column=1, value=ws_row.worker_name)
            _apply_cell_style(ws2.cell(row=row, column=1))
            ws2.cell(row=row, column=2, value=grp)
            _apply_cell_style(ws2.cell(row=row, column=2))
            ws2.cell(row=row, column=3, value=site_name)
            _apply_cell_style(ws2.cell(row=row, column=3))

            b = hours.get("基本", 0)
            o = hours.get("残業", 0)
            for col, val in [(4, b), (5, o), (6, b + o)]:
                cell = ws2.cell(row=row, column=col, value=round(val, 2))
                _apply_cell_style(cell, _HOUR_FORMAT)
            row += 1

    # 列幅
    for ws_target, widths in [
        (ws, [18, 14, 12, 12, 12, 12, 12, 15]),
        (ws2, [18, 14, 20, 12, 12, 12]),
    ]:
        for i, w in enumerate(widths, 1):
            ws_target.column_dimensions[get_column_letter(i)].width = w

    output_path = output_dir / f"個人別集計表_{result.target_month}.xlsx"
    wb.save(str(output_path))
    logger.info("個人別集計表を出力: %s", output_path)
    return output_path


def _write_group_subtotal(ws, row: int, group_name: str, cost: float, col_count: int) -> int:
    """グループ小計行を書き込み、次の行番号を返す。"""
    ws.cell(row=row, column=1, value=f"【{group_name} 小計】")
    cell = ws.cell(row=row, column=8, value=cost)
    cell.number_format = _MONEY_FORMAT
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).font = _GROUP_SUBTOTAL_FONT
        ws.cell(row=row, column=c).fill = _GROUP_SUBTOTAL_FILL
        ws.cell(row=row, column=c).border = _THIN_BORDER
    return row + 1


# ════════════════════════════════════════════
# ダッシュボードExcelエクスポート
# ════════════════════════════════════════════

_SITE_ROW_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_SITE_ROW_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
_CHILD_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_CHILD_ROW_FONT = Font(name=FONT_NAME, size=10)
_TOTAL_ROW_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
_TOTAL_ROW_FONT = Font(name=FONT_NAME, bold=True, size=11)
_OVER_BUDGET_FONT = Font(name=FONT_NAME, bold=True, size=10, color="CC0000")


def write_dashboard_export(dashboard_rows: list[dict], target_month: str = "") -> BytesIO:
    """ダッシュボードの階層データをExcelに出力し、BytesIOで返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "予算・原価消化状況"

    # タイトル行
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    month_label = f"（{target_month}）" if target_month else ""
    title_cell.value = f"予算・原価消化状況一覧{month_label}"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    # ヘッダ
    headers = ["名称", "予算額", "前回までの支払", "当月支払", "累計支払", "残金額"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _apply_header_style(ws, 3, len(headers))

    row = 4
    total_budget = 0
    total_prev = 0
    total_current = 0
    total_paid = 0
    total_remaining = 0

    for r in dashboard_rows:
        # 親行（現場合計）
        ws.cell(row=row, column=1, value=f"\u25bc {r['site_name']}")
        money_vals = [
            (2, r["budget"]), (3, r["prev_paid"]), (4, r["current_paid"]),
            (5, r["total_paid"]), (6, r["remaining"]),
        ]
        for col, val in money_vals:
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = _MONEY_FORMAT

        # 親行のスタイル
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = _SITE_ROW_FILL
            cell.font = _SITE_ROW_FONT
            cell.border = _THIN_BORDER
            if c >= 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

        if r["status"] == "進行中":
            total_budget += r["budget"]
            total_prev += r["prev_paid"]
            total_current += r["current_paid"]
            total_paid += r["total_paid"]
            total_remaining += r["remaining"]

        row += 1

        # 子行（グループ別内訳）
        for g in r.get("groups", []):
            ws.cell(row=row, column=1, value=f"  \u251c {g['group_name']}")
            g_budget = g.get("budget") or None   # 0 は未設定扱いで None に統一
            g_remaining = g.get("remaining")     # None = 未設定

            # 名称列（col=1）は上で設定済み
            # 予算額（col=2）: 0 / None は空欄
            if g_budget:
                cell2 = ws.cell(row=row, column=2, value=g_budget)
                cell2.number_format = _MONEY_FORMAT
            # 前回までの支払（col=3）
            cell3 = ws.cell(row=row, column=3, value=g["prev_paid"])
            cell3.number_format = _MONEY_FORMAT
            # 当月支払（col=4）: 必ず数値で出力（0 でも表示）
            cell4 = ws.cell(row=row, column=4, value=g["current_paid"])
            cell4.number_format = _MONEY_FORMAT
            # 累計支払（col=5）
            cell5 = ws.cell(row=row, column=5, value=g["total_paid"])
            cell5.number_format = _MONEY_FORMAT
            # 残金額（col=6）: None は空欄
            if g_remaining is not None:
                cell6 = ws.cell(row=row, column=6, value=g_remaining)
                cell6.number_format = _MONEY_FORMAT

            # 子行スタイル
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = _CHILD_ROW_FILL
                cell.font = _CHILD_ROW_FONT
                cell.border = _THIN_BORDER
                if c >= 2:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(indent=2, vertical="center")

            row += 1

    # 合計行
    ws.cell(row=row, column=1, value="合計（進行中のみ）")
    for col, val in [
        (2, total_budget), (3, total_prev), (4, total_current),
        (5, total_paid), (6, total_remaining),
    ]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = _MONEY_FORMAT

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _TOTAL_ROW_FILL
        cell.font = _TOTAL_ROW_FONT
        cell.border = _THIN_BORDER
        if c >= 2:
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # 列幅
    col_widths = [30, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行高さ = 20（データ行すべて: ヘッダ行～合計行）
    for r in range(3, row + 1):
        ws.row_dimensions[r].height = 20

    # 印刷設定
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── グループ別当月支払シート ──────────────────────────
    ws2 = wb.create_sheet("グループ別当月支払")
    month_label2 = f"（{target_month}）" if target_month else ""
    ws2.merge_cells("A1:C1")
    ws2["A1"].value = f"グループ別 当月支払一覧{month_label2}"
    ws2["A1"].font = _TITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center")

    grp_headers = ["グループ名", "現場名", "当月支払"]
    for col, h in enumerate(grp_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    _apply_header_style(ws2, 3, len(grp_headers))

    # 現場×グループの当月支払を収集しグループ順に出力
    grp_current: dict[str, list[tuple[str, float]]] = {}
    for r in dashboard_rows:
        for g in r.get("groups", []):
            grp = g["group_name"]
            if grp not in grp_current:
                grp_current[grp] = []
            grp_current[grp].append((r["site_name"], g["current_paid"]))

    r2 = 4
    grp_total_all = 0.0
    _GRP_SUBTOTAL_FILL2 = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    _GRP_SUBTOTAL_FONT2 = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")

    for grp_name, site_entries in grp_current.items():
        grp_subtotal = 0.0
        for site_name, current_paid in site_entries:
            ws2.cell(row=r2, column=1, value=grp_name)
            _apply_cell_style(ws2.cell(row=r2, column=1))
            ws2.cell(row=r2, column=2, value=site_name)
            _apply_cell_style(ws2.cell(row=r2, column=2))
            cell = ws2.cell(row=r2, column=3, value=current_paid)
            _apply_cell_style(cell, _MONEY_FORMAT)
            grp_subtotal += current_paid
            r2 += 1

        # グループ小計行
        ws2.cell(row=r2, column=1, value=f"【{grp_name} 小計】")
        ws2.cell(row=r2, column=2, value="")
        cell = ws2.cell(row=r2, column=3, value=grp_subtotal)
        cell.number_format = _MONEY_FORMAT
        for c in range(1, 4):
            ws2.cell(row=r2, column=c).font = _GRP_SUBTOTAL_FONT2
            ws2.cell(row=r2, column=c).fill = _GRP_SUBTOTAL_FILL2
            ws2.cell(row=r2, column=c).border = _THIN_BORDER
        grp_total_all += grp_subtotal
        r2 += 1

    # 総計行
    ws2.cell(row=r2, column=1, value="総計")
    ws2.cell(row=r2, column=2, value="")
    cell = ws2.cell(row=r2, column=3, value=grp_total_all)
    cell.number_format = _MONEY_FORMAT
    for c in range(1, 4):
        ws2.cell(row=r2, column=c).fill = _TOTAL_ROW_FILL
        ws2.cell(row=r2, column=c).font = _TOTAL_ROW_FONT
        ws2.cell(row=r2, column=c).border = _THIN_BORDER

    for i, w in enumerate([20, 28, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for rr in range(3, r2 + 1):
        ws2.row_dimensions[rr].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("ダッシュボードExcelを生成しました")
    return buf
