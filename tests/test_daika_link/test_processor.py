"""skills/daika_link_setting/processor.py のユニットテスト。

純粋関数（normalize_key, extract_keyword）と Excel I/O を伴う
インテグレーション関数（process, _build_target_index, _apply_links）を、
合成した小さな .xlsx ファイルで検証する。
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from skills.daika_link_setting import process
from skills.daika_link_setting.processor import (
    LinkStats,
    _build_target_index,
    _find_basis_column,
    _resolve_anchor,
    extract_keyword,
    normalize_key,
)


# ────────────────────────────────────────────
# normalize_key
# ────────────────────────────────────────────
class TestNormalizeKey:
    def test_none(self):
        assert normalize_key(None) == ""

    def test_empty(self):
        assert normalize_key("") == ""

    @pytest.mark.parametrize("value,expected", [
        ("単 1 号", "単1号"),
        ("単　1　号", "単1号"),         # 全角空白
        ("単１号", "単1号"),            # 全角数字 → NFKC
        ("Ｐ1号", "P1号"),              # 全角 Ｐ → P
        ("-0001 Ｐ 1 号", "-0001P1号"),
        ("-0001 施 1 号", "-0001施1号"),
        ("明  3  号", "明3号"),
        ("W", "W"),
        ("10", "10"),
    ])
    def test_examples(self, value, expected):
        assert normalize_key(value) == expected

    def test_int_passthrough(self):
        assert normalize_key(10) == "10"


# ────────────────────────────────────────────
# extract_keyword
# ────────────────────────────────────────────
class TestExtractKeyword:
    @pytest.mark.parametrize("normalized,expected", [
        ("単1号", "単1号"),
        ("施12号", "施12号"),
        ("P1号", "P1号"),
        ("明3号", "明3号"),
        ("-0001P1号", "P1号"),
        ("-0001施1号", "施1号"),
    ])
    def test_match(self, normalized, expected):
        assert extract_keyword(normalized) == expected

    @pytest.mark.parametrize("normalized", ["", "W", "X", "10", "基準", "号", "単号"])
    def test_no_match(self, normalized):
        assert extract_keyword(normalized) is None


# ────────────────────────────────────────────
# Excel fixture
# ────────────────────────────────────────────
def _build_sample_workbook() -> BytesIO:
    """簡略化したテスト用 xlsx を BytesIO に生成する。

    ・「本工事費表」: M列に「基準」ヘッダ＋数行のキーワード
    ・「特別単価表A」: C列の結合セル C4:D4 に「単1号」
    ・「代価表」: C列に「施1号」
    ・「施工パッケージ代価表」: A列に「P1号」
    """
    wb = Workbook()
    # デフォルトシートを「本工事費表」にリネーム
    ws_main = wb.active
    ws_main.title = "本工事費表"
    # M1 に「基準」ヘッダ
    ws_main.cell(row=1, column=13, value="基　　　準")
    # M列にキーワード（一部はターゲットなし／プレフィックスあり）
    ws_main.cell(row=2, column=13, value="単 1 号")
    ws_main.cell(row=3, column=13, value="-0001 施 1 号")
    ws_main.cell(row=4, column=13, value="-0001 Ｐ 1 号")
    ws_main.cell(row=5, column=13, value="W")              # リンクなし
    ws_main.cell(row=6, column=13, value="単 99 号")       # ターゲットなし → unmatched
    ws_main.cell(row=7, column=13, value="明 1 号")        # ターゲットなし → unmatched

    # 特別単価表A: C4:D4 結合セルに「単1号」
    ws_tan = wb.create_sheet("特別単価表A")
    ws_tan.cell(row=1, column=15, value="基準")  # O列ヘッダ
    ws_tan.cell(row=4, column=3, value="単1号")
    ws_tan.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4)

    # 代価表: C列に「施1号」（非結合）
    ws_daika = wb.create_sheet("代価表")
    ws_daika.cell(row=1, column=15, value="基準")
    ws_daika.cell(row=10, column=3, value="施1号")

    # 施工パッケージ代価表: A列に「P1号」
    ws_pkg = wb.create_sheet("施工パッケージ代価表")
    ws_pkg.cell(row=1, column=1, value="P1号")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def sample_workbook_bytes() -> bytes:
    return _build_sample_workbook().getvalue()


# ────────────────────────────────────────────
# _find_basis_column / _resolve_anchor
# ────────────────────────────────────────────
class TestSheetHelpers:
    def test_find_basis_column_main(self, sample_workbook_bytes):
        wb = load_workbook(BytesIO(sample_workbook_bytes))
        ws = wb["本工事費表"]
        assert _find_basis_column(ws) == 13  # M列

    def test_find_basis_column_missing(self, sample_workbook_bytes):
        wb = load_workbook(BytesIO(sample_workbook_bytes))
        ws = wb["施工パッケージ代価表"]
        assert _find_basis_column(ws) is None

    def test_resolve_anchor_for_merged_cell(self, sample_workbook_bytes):
        wb = load_workbook(BytesIO(sample_workbook_bytes))
        ws = wb["特別単価表A"]
        # 結合範囲の右下セル D4 を渡しても左上 C4 が返る
        d4 = ws.cell(row=4, column=4)
        assert _resolve_anchor(ws, d4) == "C4"

    def test_resolve_anchor_for_non_merged(self, sample_workbook_bytes):
        wb = load_workbook(BytesIO(sample_workbook_bytes))
        ws = wb["代価表"]
        cell = ws.cell(row=10, column=3)
        assert _resolve_anchor(ws, cell) == "C10"


# ────────────────────────────────────────────
# _build_target_index
# ────────────────────────────────────────────
class TestBuildTargetIndex:
    def test_collects_all_targets(self, sample_workbook_bytes):
        wb = load_workbook(BytesIO(sample_workbook_bytes))
        stats = LinkStats()
        index = _build_target_index(wb, stats)
        assert index["単1号"] == ("特別単価表A", "C4")
        assert index["施1号"] == ("代価表", "C10")
        assert index["P1号"] == ("施工パッケージ代価表", "A1")
        # 明1号 はターゲットがないのでインデックスに無い
        assert "明1号" not in index
        # 単99号 も無い
        assert "単99号" not in index
        assert stats.duplicate_targets == []


# ────────────────────────────────────────────
# process（end-to-end）
# ────────────────────────────────────────────
class TestProcess:
    def test_links_are_applied(self, sample_workbook_bytes, tmp_path: Path):
        out_path = tmp_path / "out.xlsx"
        stats = process(sample_workbook_bytes, out_path)

        # 単/施/P の 3 件はリンク化、明1号と単99号は未マッチ、W は対象外
        assert stats.links_created == 3
        # 4 シート全てを走査
        assert stats.sheets_scanned == 4
        # ターゲットは 3 個（単1号 / 施1号 / P1号）
        assert stats.targets_indexed == 3
        # 未マッチは 単99号 / 明1号 の 2 件
        assert set(stats.keywords_unmatched) == {"単99号", "明1号"}
        # 基準列マップ
        assert stats.basis_columns.get("本工事費表") == "M"

        # 出力ファイルを再読み込みしてハイパーリンクが付与されていることを確認
        wb = load_workbook(out_path)
        ws = wb["本工事費表"]
        # M2「単 1 号」→ 特別単価表A!C4
        cell_tan = ws.cell(row=2, column=13)
        assert cell_tan.hyperlink is not None
        assert "特別単価表A" in cell_tan.hyperlink.location
        assert "C4" in cell_tan.hyperlink.location
        # M3「-0001 施 1 号」→ 代価表!C10
        cell_se = ws.cell(row=3, column=13)
        assert cell_se.hyperlink is not None
        assert "代価表" in cell_se.hyperlink.location
        assert "C10" in cell_se.hyperlink.location
        # M4「-0001 Ｐ 1 号」→ 施工パッケージ代価表!A1
        cell_p = ws.cell(row=4, column=13)
        assert cell_p.hyperlink is not None
        assert "施工パッケージ代価表" in cell_p.hyperlink.location
        assert "A1" in cell_p.hyperlink.location
        # M5 (W) はリンクなし
        assert ws.cell(row=5, column=13).hyperlink is None
        # M6 (単 99 号) は未マッチなのでリンクなし
        assert ws.cell(row=6, column=13).hyperlink is None
        # 単1号セル自身（特別単価表A C4）はリンクなし
        wb2 = load_workbook(out_path)
        assert wb2["特別単価表A"].cell(row=4, column=3).hyperlink is None

    def test_accepts_path_input(self, sample_workbook_bytes, tmp_path: Path):
        # Path 経由でも動作することを確認
        src_path = tmp_path / "src.xlsx"
        src_path.write_bytes(sample_workbook_bytes)
        out_path = tmp_path / "out.xlsx"
        stats = process(src_path, out_path)
        assert stats.links_created == 3


# ────────────────────────────────────────────
# 重複ターゲット
# ────────────────────────────────────────────
class TestDuplicateTargets:
    """同名キーが複数シートに現れた場合、最初の出現位置を採用し
    重複は duplicate_targets に記録されることを確認する。"""

    def test_duplicate_recorded(self, tmp_path: Path):
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "本工事費表"
        ws_main.cell(row=1, column=13, value="基準")
        ws_main.cell(row=2, column=13, value="単 1 号")

        # 2 つのシートに「単1号」を持たせる
        ws_a = wb.create_sheet("代価表")
        ws_a.cell(row=5, column=3, value="単1号")
        ws_b = wb.create_sheet("特別単価表A")
        ws_b.cell(row=10, column=3, value="単1号")

        src = BytesIO()
        wb.save(src)
        out_path = tmp_path / "out.xlsx"

        stats = process(src.getvalue(), out_path)
        # シート順「代価表」が先、「特別単価表A」が後 → 代価表が採用される
        assert stats.links_created == 1
        assert stats.duplicate_targets == ["単1号"]

        wb2 = load_workbook(out_path)
        link = wb2["本工事費表"].cell(row=2, column=13).hyperlink
        assert link is not None
        assert "代価表" in link.location
